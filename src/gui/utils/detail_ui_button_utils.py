# -*- coding: utf-8 -*-
# @Time    : 2025/4/26 1:04
# @Author  : ESJIAN
# @Email   : esjian@outlook.com
# @File    : ui_utils.py
# @Software: VsCode




from datetime import datetime

from PySide6.QtWidgets import QApplication, QWidget
import pandas as pd  
import os
import xlrd# 用于获取所有子表sheet名

#这两个用来操作识别完成的暂存表修改数据
from openpyxl import load_workbook
import Levenshtein

from configparser import ConfigParser
from PySide6.QtWidgets import QVBoxLayout, QLabel, QPushButton, QLineEdit,QHBoxLayout,QGroupBox

from src.gui.error_window import TagNumShortage,IndexOutOfRange                 # Learning1：子模块的导入相对路径的起算点是主模块
from src.gui.check_window import ExcelCheckWindow               # Learning2:顶级脚本设定绝对倒入配置后不需要在子模块中重设
from src.gui.data_save_dialog import data_save_success
from src.core.excel_handler import store_single_entry_to_temple_excel  # Fixed1:将项目包以绝对形式导入,解决了相对导入不支持父包的报错
from src.core.excel_handler import commit_data_to_storage_excel
import xlwings as xl

import __main__                                                 # Learning5:__main__模块的引用，访问主模块变量



def get_current_date():
    """
    获取系统当前日期，格式为 YYYY-MM-DD
    :param: None
    :return: 当前日期的字符串，格式为 YYYY-MM-DD
    """
    return datetime.now().strftime("%Y-%m-%d")


def manual_temp_storage(self,input_fields):
    """
    暂存当前条目所有输入框内的信息
    :param input_fields: 输入框的字典或列表，键为字段名，值为对应的 QLineEdit 对象
    :return: 包含所有输入框内容的字典
    """
    values_must_have = [
        self.line1Right,  # 日期
        self.line2Right,  # 品名
        self.line3Right,  # 类别
        self.line6Right,  # 数量
        self.line5Right,  # 金额
        self.line9Right,  # 公司
        self.line10Right  # 单名
    ]

    for i in values_must_have:
        if i.text() == "":
            i.setPlaceholderText("该项必填")
            #显示错误窗口
            show_error_window(self)
            return None


    __main__.TEMP_LIST_ROLLBACK_SIGNAL = False  # type: ignore # Learning3：信号量，标记是否需要回滚
    exsit_tag_number = 0                        # 统计有内容的输入框数量

    temp_storage = {}                           # 存储输入框内容的字典

    try:
        for field_name, input_field in input_fields.items():
            if __main__.DEBUG_SIGN == True:
                exsit_tag_number+=1
                temp_storage[field_name] = input_field

            elif input_field.text() :              # 检查输入框是否有内容
                exsit_tag_number+=1                # 统计有内容的输入框数量
                temp_storage[field_name] = input_field.text() # 将输入框内容存储到字典
        if __main__.DEBUG_SIGN == True:
            __main__.SERIALS_NUMBER += 1   # type: ignore

        if exsit_tag_number==__main__.TOTAL_FIELD_NUMBER:

            self.line1Right.setText("")         # Learning4：对QLineEdit组件使用setText()方法重置输入框内容
            self.line2Right.setText("")
            self.line3Right.setText("")
            self.line4Right.setText("")
            self.line5Right.setText("")
            self.line6Right.setText("")
            self.line7Right.setText("")
            self.line8Right.setText("")
            self.line9Right.setText("")
            self.line10Right.setText("")

            "重置输入框提示文本为上一次的输入"
            self.line1Right.setPlaceholderText(input_fields['日期'])
            self.line2Right.setPlaceholderText(input_fields['类别'])
            self.line3Right.setPlaceholderText(input_fields['品名'])
            self.line8Right.setPlaceholderText(input_fields['单位'])
            self.line7Right.setPlaceholderText(input_fields['单价'])
            self.line6Right.setPlaceholderText(input_fields['数量'])
            self.line5Right.setPlaceholderText(input_fields['金额'])
            self.line4Right.setPlaceholderText(input_fields['备注'])
            self.line9Right.setPlaceholderText(input_fields['公司'])
            self.line10Right.setPlaceholderText(input_fields['单名'])
            
            
            "更新信息栏信息"
            try:
                self.storageNum.setText(str(__main__.TEMP_STORAGED_NUMBER_LISTS))  # 更新存储数量的标签文本
                self.spinBox.setValue(__main__.TEMP_STORAGED_NUMBER_LISTS+1)           # 更新正在编辑第 xx 项目的 xx 数值
                __main__.TEMP_STORAGED_NUMBER_LISTS +=1                              # type: ignore # Learning5：形式参数传参进来的变量
            except Exception as e:
                print(f"Error: {e}")
                return None
            
            print("Notice: 暂存数据为", temp_storage)
            
            # 调用 store_single_entry_to_excel 函数存储数据到Excel文件,以xls方式存储
            store_single_entry_to_temple_excel(self,temp_storage, __main__.TEMP_SINGLE_STORAGE_EXCEL_PATH)
            
            # 暂存列表展示回滚                       
            temp_list_rollback(self)
            return temp_storage
        
        else:
            print("Warning: Not all fields are filled.")
            show_error_window(self) # 显示错误窗口
            return None
    except Exception as e:
        print(f"Error: {e}")
        return None
    

def show_error_window(self):
    """
    显示错误窗口
    :param: None
    :return: None
    """
    
    if hasattr(self, 'Form') and self.PopWindowApplicationForm.isVisible():
        return  # 如果弹窗已经存在且可见，则不重复创建
    
    # 检验self是否有名为app的属性
    self.PopWindowApplication = QApplication.instance()
    if not self.PopWindowApplication:
        # 若没有为self追加创建一个app属性,继承自QApplication
        self.PopWindowApplication = QApplication([])
        # 为self追加创建一个Form属性,继承自QWidget
        self.PopWindowApplicationForm = QWidget()
        # 为self追加一个ui属性,继承自TagNumShortage
        self.PopWindowApplicationUi = TagNumShortage()
        #
        self.PopWindowApplicationUi.setupUi(self.PopWindowApplicationForm)
        
        self.PopWindowApplicationForm.show()
    else:
        # 为self追加创建一个Form属性,继承自QWidget
        self.PopWindowApplicationForm = QWidget()
        # 为self追加一个ui属性,继承自TagNumShortage
        self.PopWindowApplicationUi = TagNumShortage()
        #
        self.PopWindowApplicationUi.setupUi(self.PopWindowApplicationForm)
        
        self.PopWindowApplicationForm.show()


def show_check_window(self,file_path):
    """
    显示检查窗口
    :param: self,file_path
    :return: None
    """
    # 为self追加创建一个app属性,继承自QApplication
    #self.PopWindowAplication = QApplication([])  # Learning3&Fixed2:一个程序只能有一个 QApplication 实例


    # 为self追加创建一个Form属性,继承自QWidget
    self.PopWindowApplicationForm = QWidget()
    # 为self追加一个ui属性,继承自excel_check_window
    self.PopWindowApplicationUi = ExcelCheckWindow()
    # 将 ui 属性与 form 属性 
    self.PopWindowApplicationUi.set_up_Ui(self.PopWindowApplicationForm)
    
    self.PopWindowApplicationUi.load_table_data(file_path)

    # 展示窗口
    self.PopWindowApplicationForm.show()
    

def commit_data_to_excel(self,model,main_excel_file_path,sub_main_food_workbook,sub_auxiliary_food_workbook,welfare_food_workbook):
    """
    提交数据到主表、副表Excel文件

    Parameters:
     - self: Object    
     - model: 输入输出模式切换变量,0为入库模式/1为出库模式
     - excel_file_path: 主表Excel文件路径
     - sub_main_food_workbook: 子食主食表Excel文件路径
     - sub_auxiliary_food_workbook: 子食副食表Excel文件路径
     - welfare_food_workbook: 福利副表Excel文件路径
    
    Return: None
    """

    with xlrd.open_workbook(__main__.TEMP_SINGLE_STORAGE_EXCEL_PATH) as read_temp_storage_workbook:

        "判断表是否为空"
        if read_temp_storage_workbook.sheet_by_index(0).nrows == 1:
            print(f"Warning: 暂存表 {__main__.TEMP_SINGLE_STORAGE_EXCEL_PATH} 为空,请先添加数据")
            # 弹窗提示暂存表为空
            self.worker.signal3.emit()
            self.pushButton_5.setText("提交数据")
            return
    
    

    commit_data_to_storage_excel(self,model,main_excel_file_path,sub_main_food_workbook,sub_auxiliary_food_workbook,welfare_food_workbook)


def temp_list_rollback(self):
    """
    实现点击信息栏中“正在编辑xx项目”上下箭头时,正在编辑条目回滚的视图回滚
    :param: self
    :return: None
    """

    print(f"Notice:当前编辑条目为第{self.spinBox.value()}项,条目切换信号为{__main__.TEMP_LIST_ROLLBACK_SIGNAL}")

    if self.spinBox.value()>0 and __main__.TEMP_LIST_ROLLBACK_SIGNAL == True: # Learning6：py的与符号是and关键字而不是&，&是位运算符
        
        try:
            # 如果目标表格不存在则调用xlwings创建一份 TEMP_SINGLE_STORAGE_EXCEL_PATH 文件
            if not os.path.exists(__main__.TEMP_SINGLE_STORAGE_EXCEL_PATH): # Learning4：判断文件是否存在
                # 创建一个空的Excel文件
                app = xl.App(visible=False)
                workbook = xl.Book()
                sheet = workbook.sheets[0]
                # 为sheet  添加表头 ['日期', '类别', '品名', '单位', '单价', '数量', '金额', '备注', '公司','单名']
                sheet.range('A1').value = ['日期', '类别', '品名', '单位', '单价', '数量', '金额', '备注', '公司','单名']
                workbook.save(__main__.TEMP_SINGLE_STORAGE_EXCEL_PATH)
                workbook.close()
                app.quit()
                print(f"Warning: temp_manual_input_data.xls 不存在，已自动创建")
            
            temp_storage = pd.read_excel(__main__.TEMP_SINGLE_STORAGE_EXCEL_PATH)
            print(f"Notice: {temp_storage}")
            index = self.spinBox.value()
            
    
            # 如果目标条目索引号在已存储列表范围内，则切换到阅览已存储条目模式
            if 0 < index <= len(temp_storage):#这里左边改成开区间了, 不能为0, 下同
                # 获取当前条目的数据
                current_entry = temp_storage.iloc[index-1]
                # 更新输入框内容
                self.line1Right.setText(str(current_entry['日期']))
                self.line2Right.setText(str(current_entry['类别']))
                self.line3Right.setText(str(current_entry['品名']))
                self.line4Right.setText(str(current_entry['备注']))
                self.line5Right.setText(str(current_entry['金额']))
                self.line6Right.setText(str(current_entry['数量']))
                self.line7Right.setText(str(current_entry['单价']))
                self.line8Right.setText(str(current_entry['单位']))
                self.line9Right.setText(str(current_entry['公司']))
                self.line10Right.setText(str(current_entry['单名']))

            # 如果目标条目索引号超出已存储列表+1，则切换到输入条目模式
            elif 0 < index <= len(temp_storage)+1:
                self.line1Right.setText("")
                self.line2Right.setText("")
                self.line3Right.setText("")
                self.line4Right.setText("")
                self.line5Right.setText("")
                self.line6Right.setText("")
                self.line7Right.setText("")
                self.line8Right.setText("")
                self.line9Right.setText("")
                self.line10Right.setText("")

            # 如果目标编辑条目索引号超出已存储列表范围+1，则提示错误，并且返回最后改动的条目上
            else:
                print("Notice: 条目超出范围，请检查条目索引号")
                # 重置条目索引到报错前
                self.spinBox.setValue(__main__.TEMP_STORAGED_NUMBER_LISTS) # 更新SpinBox的值为存储数量
                # 弹窗报错
                # 为self追加创建一个Form属性,继承自QWidget
                self.PopWindowApplicationForm = QWidget()
                # 为self追加一个ui属性,继承自TagNumShortage
                self.PopWindowApplicationUi = IndexOutOfRange() # Learning7：不要误用成self.IndexOutOfRange(self)，因为IndexOutOfRange是一个类，而不是一个函数
                self.PopWindowApplicationUi.setupUi(self.PopWindowApplicationForm)     
                self.PopWindowApplicationForm.show()
                return None
                
        except Exception as e:
                    print(f"Error: {e}")
                    return None
    else:
        #将其设置为1
        if self.spinBox.value() <= 0:
            self.spinBox.setValue(1)
        __main__.TEMP_LIST_ROLLBACK_SIGNAL = True
        
def show_setting_window(self):
    """
    显示设置窗口
    :param: self
    :return: None
    """
    self.settings_window = QWidget()
    self.settings_window.setWindowTitle("设置")
    self.settings_window.resize(400, 300)

    # 将此布局与父widget关联
    main_layout = QHBoxLayout(self.settings_window)
    
    # 创建子Group Box 布局组件
    self.manual_group_box = QGroupBox(self.settings_window)
    self.manual_group_box.setTitle("手动导入设置")
    self.img_group_box = QGroupBox(self.settings_window)
    self.img_group_box.setTitle("照片导入设置")
    # 为 setting_window 对两个group box 设置布局
    main_layout.addWidget(self.manual_group_box)
    main_layout.addWidget(self.img_group_box)

    

    # 设定手动导入管理选项
    manual_layout = QVBoxLayout(self.manual_group_box)

    # Add toggle for "Auto-fill Date"
    self.auto_fill_date_toggle = QPushButton("自动填充日期")
    self.auto_fill_date_toggle.setCheckable(True)
    self.auto_fill_date_toggle.setChecked(get_ini_setting("Settings", "auto_fill_date", file_path='../../config/config.ini') == 'True')
    print(get_ini_setting("Settings", "auto_fill_date", file_path='../../config.ini') == 'True')
    self.auto_fill_date_toggle.clicked.connect(lambda: modify_ini_setting("Settings", "auto_fill_date", self.auto_fill_date_toggle.isChecked()))
    manual_layout.addWidget(self.auto_fill_date_toggle)

    # Add toggle for "Auto-calculate Total Price"
    self.auto_calc_price_toggle = QPushButton("自动根据单价数量计算总价")
    self.auto_calc_price_toggle.setCheckable(True)
    self.auto_calc_price_toggle.setChecked(get_ini_setting("Settings", "auto_calc_price", file_path='../../config/config.ini') == 'True')
    self.auto_calc_price_toggle.clicked.connect(lambda: modify_ini_setting("Settings", "auto_calc_price", self.auto_calc_price_toggle.isChecked()))
    manual_layout.addWidget(self.auto_calc_price_toggle)

    close_button = QPushButton("点击关闭")
    close_button.clicked.connect(self.settings_window.close)
    manual_layout.addWidget(close_button)

    # 设定照片导入管理选项
    img_layout = QVBoxLayout(self.img_group_box)

    # Add toggle for "Auto-fill Date"
    self.img_auto_fill_date_toggle = QPushButton("自动填充日期")
    self.img_auto_fill_date_toggle.setCheckable(True)
    self.img_auto_fill_date_toggle.setChecked(get_ini_setting("Settings", "img_auto_fill_date", file_path='../../config/config.ini') == 'True')
    #print(get_ini_setting("Settings", "img_auto_fill_date", file_path='../../config.ini') == 'True')
    self.img_auto_fill_date_toggle.clicked.connect(lambda: modify_ini_setting("Settings", "img_auto_fill_date", self.img_auto_fill_date_toggle.isChecked()))
    

    # Add toggle for "Auto-calculate Total Price"
    self.img_auto_calc_price_toggle = QPushButton("自动根据单价数量计算总价")
    self.img_auto_calc_price_toggle.setCheckable(True)
    self.img_auto_calc_price_toggle.setChecked(get_ini_setting("Settings", "img_auto_calc_price", file_path='../../config/config.ini') == 'True')
    self.img_auto_calc_price_toggle.clicked.connect(lambda: modify_ini_setting("Settings", "img_auto_calc_price", self.img_auto_calc_price_toggle.isChecked()))
    

    close_button = QPushButton("点击关闭")
    close_button.clicked.connect(self.settings_window.close)

    # 将子组件放入布局组件进行管理
    img_layout.addWidget(self.img_auto_fill_date_toggle)
    img_layout.addWidget(self.img_auto_calc_price_toggle)
    img_layout.addWidget(close_button)

    
    # 设定主窗口的布局
    self.settings_window.setLayout(main_layout)
    # 显示窗口
    self.settings_window.show()

def close_setting_window(self):
    """
    关闭设置窗口
    :param: self
    :return: None
    """
    if hasattr(self, 'settings_window'):
        self.settings_window.close()
        del self.settings_window

def modify_ini_setting(section, option, new_value, file_path='../../config/config.ini'):
    """
    修改INI配置文件中指定的设置项。如果文件、段或选项不存在则自动创建。

    参数:
        section (str): 配置段名
        option (str): 配置项名
        new_value (str): 新的值
        file_path (str): INI 文件路径

    返回:
        bool: 修改成功返回 True，失败返回 False
    """
    config = ConfigParser()
    # 如果文件存在就读取；否则创建空文件
    new_value = str(new_value)   
    if os.path.isfile(file_path):
        config.read(file_path, encoding='utf-8')
    else:
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("")  # 创建空文件

    try:
        if section not in config:
            config.add_section(section)

        config[section][option] = new_value

        with open(file_path, 'w', encoding='utf-8') as f:
            config.write(f)
        print(f"'{section}' 中 '{option}' 的值已修改为: {new_value}")
        return True
    except Exception as e:
        print(f"Error: {e}")
        print(f"Error: 无法修改 '{section}' 中 '{option}' 的值")
        return False

def get_ini_setting(section, option, file_path='../../config/config.ini'):
    """
    从INI配置文件中获取指定设置项的值。
    若文件、段或选项不存在，则自动创建并写入 "False"，然后返回 "False"。

    参数:
        section (str): 配置段名
        option (str): 配置项名
        file_path (str): INI 文件路径

    返回:
        str: 配置项的值；若不存在则返回 "False"
    """
    config = ConfigParser()

    # 文件不存在：创建空文件
    if not os.path.isfile(file_path):
        os.makedirs(os.path.dirname(file_path), exist_ok=True) if os.path.dirname(file_path) else None
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write("")

    config.read(file_path, encoding='utf-8')

    updated = False

    if section not in config:
        config.add_section(section)
        updated = True

    if option not in config[section]:
        config[section][option] = "False"
        updated = True

    if updated:
        with open(file_path, 'w', encoding='utf-8') as f:
            config.write(f)
        return "False"

    return config.get(section, option)

def convert_place_holder_to_text(self):
    """
    将当前输入框的占位符转换为文本
    :param: self
    :return: None
    """
    #检查当前聚焦的输入框
    current_widget = self.focusWidget()
    if isinstance(current_widget, QLineEdit):
        #获取当前输入框的占位符文本
        placeholder_text = current_widget.placeholderText()
        #将占位符文本设置为输入框的文本
        current_widget.setText(placeholder_text)
    
def cancel_input_focus(self):
    """
    取消当前输入框的焦点
    :param: self
    :return: None
    """
    #检查当前聚焦的输入框
    current_widget = self.focusWidget()
    if isinstance(current_widget, QLineEdit):
        #取消当前输入框的焦点
        current_widget.clearFocus()

def mode_not_right(self, MODE):
    text = self.line10Right.text()
    if "入库" in text or "出库" in text:
        # MODE == 0 表示当前系统是入库模式
        if "出库" in text and MODE == 0:
            return True  # 模式是入库，但写了出库
        elif "入库" in text and MODE == 1:
            return True  # 模式是出库，但写了入库
    return False

def fetch_all_product_names():
    # 获取当前目录下所有 .xls 文件
    xls_files = [os.path.join(".", "src", "data", "storage", "work", "子表", f) for f in os.listdir(os.path.join(".", "src", "data", "storage", "work", "子表")) if f.endswith('.xls')]
    print(xls_files)
    res = []
    for file in xls_files:
        try:
            workbook = xlrd.open_workbook(file)
            sheet_names = workbook.sheet_names()
            res += sheet_names
            # print(f"{file} 的 sheets: {sheet_names}")
        except Exception as e:
            print(f"{file} 打开失败: {e}")
    return res

def modify_data_in_image_excel(self):
    """
    修正图像识别结果表格的品名
    """
    print("开始修正图像识别结果表格数据")
    # 从子表提取的所有品名，其实这个应该动态获取两个子表的sheet名
    # products = ['排骨15', '前槽', '精五花', '里脊', '精肉', '叉骨', '猪蹄', '猪肝', '猪头肉1', '脊骨', '骨棒', '肘子', '腿骨', '大骨棒', '筋蹄', '护心肉', '牛腩', '牛腱子', '羊排', '蛋鸡', '牛肉', '鸡腿', '鸡胸肉', '护心肉1', '正大鸡翅根', '鸡翅中', '鸡全翅', '鸡手', '鸭腿', '鸭子', '大鹅', '黄花鱼', '鳕鱼 ', '带鱼  (2)', '鲤鱼', '青鱼', '鲫鱼', '鳕鱼', '草鱼 ', '鱿鱼', '虾', '虾仁', '绿豆芽1', '金针菇', '大豆腐', '豆干+豆腐干', '冻豆腐 (2)', '干豆腐', '香菇', '豆干', '娃娃菜', '干榛蘑', '鲜蘑', '大辣椒', '青椒', '尖椒', '辣妹子', '香菇1', '蒜薹', '莴笋', '地瓜', '杏鲍菇', '白菜', '西红柿', '菠菜', '油菜 ', '生菜  ', '胡萝卜10', '绿萝卜', '萝卜', '油麦菜1', '苦苣', '甘蓝', '黄瓜', '蒜苔', '西兰花', '茄子', '土豆52', '黄豆芽', '苦瓜', '蘑菇', '山药', '韭菜', '豆角', '冬瓜', '芹菜', '小白菜', '菜花', '油麦菜', '倭瓜窝瓜', '紫薯', '角瓜', '豆芽类', '藕', '鸡蛋24', '葱15', '姜3', '蒜4', '圆葱20', '香菜', '紫菜3', '海带结', '酸菜20', '腐竹', '鸡丸', '粉丝', '粉条6', '盐15', '白糖7', '红糖 (2)', '料酒8', '陈醋1', '芝麻酱', '香醋', '蚝油', '奥尔良烤翅腌', '香油（芝麻油）1', '酱油1', '红烧酱料（老抽）', '捞汁', '豆油', '鸡精', '十三香3', '桂皮8', '淀粉（生粉）', '胡椒粉2', '大料14', '花椒7', '香叶16', '孜然', '芝麻', '牛肉粉3', '辣椒粉', '辣椒段18', '辣椒片', '花生米4', '麻婆豆腐料', '水煮肉片料', '腐竹1', '银耳', '木耳', '红九九', '火锅底料', '蒜蓉酱', '酸黄瓜', '麻辣鲜', '郫县豆瓣酱3', '黄豆酱13', '火腿', '午餐肉', '松花蛋', '火腿肠', '方便面 (104g)', '方便面桶（110g）', '纯牛奶', '榨菜', '八宝粥', '月饼', '番茄酱', '白玉菇', '猪头肉', '咖喱', '辣白菜', '腐乳', '白醋', '咸鸭蛋', '冰糖10', '粽子', '香蕉1', '酸奶', '蜜瓜', '香梨', '苹果1', '圣女果', '西瓜 (2)', '火龙果', '油（扶贫）', '菜籽油（扶贫）', '花菇（扶贫）', '香菇（扶贫）', '苹果', '香蕉', '梨', '桃', '葡萄', '桔子', '扶贫黑木耳 (2)', '扶贫围场黑木耳', '扶贫木耳', '红薯粉', '食诺土豆粉 (2)', '糖三角', '火烧饼 ', '厂玉米面馒头', '厂月饼2)', '厂甜红豆卷', '厂红枣香酥包)', '厂豆包', '厂豆沙包 ', '豆沙饼', '厂无水蛋糕', '厂糖饼', '厂发糕', '厂麻花 ', '厂烤饼', '厂包子', '厂馒头100', '厂花卷', '黑米', '大米', '面 粉(2)', '小米', '红小豆', '饭豆', '玉米面', '小碴子', '燕麦米', '绿豆', '花豆 ', '挂面15', '手擀面', '帮扶小米', '扶贫大米61', '扶贫面粉（不入账）', '面粉（帮扶）', '豆包（帮扶', '玉米（扶贫）', '米粉（扶贫）', '扶贫挂面', '馒头', '花卷', '烤饼', '大馇子', '饭豆1', 'Sheet1']
    products = fetch_all_product_names()
    # 打开 Excel 文件
    save_folder = "./src/data/input/manual"
    excel_path = os.path.join(save_folder, f"temp_img_input.xlsx")
    wb = load_workbook(excel_path)
    ws = wb.active# 默认sheet
    # 遍历第 C 列（即第 3 列）从第 2 行开始
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        cell = row[0]
        if cell.value is not None and str(cell.value).strip() != "":
            image_product = cell.value
            ratios = dict()  # 每个标准名对应一个近似度，取近似度最大的那个名字
            for product_name in products:
                ratios[product_name] = Levenshtein.ratio(image_product, product_name)
            # 取相似度最大的词
            best_name = max(ratios, key=ratios.get)
            print(f"{cell.value} -- > {best_name}")
            ws.cell(row=cell.row, column=2, value=best_name)

    wb.save(excel_path)




# Summary：
# 1. 抽象的看所有widget都是一个个的对象，都是有属性和方法的，属性就是它的状态，方法就是它能做的事情。


# Learning：
# 1. 弹窗事件循环的生命周期依赖于该调用函数的声明周期，如果该函数是一个独立的函数，那么它的生命周期就会在函数调用结束后结束，导致弹窗无法正常显示。
#    子弹窗要想和共享主窗口的声明周期，需要利用self的引用来实现。也就是在主窗口中创建子窗口的实例，并将其作为属性存储在主窗口中，这样子窗口就可以和主窗口共享生命周期了。
#    来源：Fixed1的修复
# 2. 顶级脚本设定绝对导入包配置代码后不需要在子模块中重设，因为主代码设置好的变量子代码可见
#    变量的传递分为单个文件内，和跨文件两种方式去传递
# 3. 在存在主窗口的时候，在创建子窗口时，不要创建新的QApplication实例，而是使用已经存在的实例。因为QApplication只能有一个实例，创建多个实例会导致错误。
#    来源：Fixed2的修复
# 4. 如果直接是self.date_2 = ""，self.date_2不再指向原来的 QLineEdit 对象，而是被重新赋值为一个字符串 ""
# 5. 
# 6. 
# 7. 明确一个类你是想调用还是想实例化，调用就是直接使用类名加括号，实例化就是以创建一个对象。
