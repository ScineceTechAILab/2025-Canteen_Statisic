"""
item_data_operate.py: Utility functions for item inventory operations.

This module provides a collection of functions for updating and retrieving
item information in the inventory index during item check-in and check-out
operations.

Author: ESJIAN
Email: esjian@outlook.com

Copyright (c) 2025 STA LAB. All rights reserved.
Licensed under the MIT License.

Version: 1.0
Date: 2025-7-6

Dependencies:
    - [List any dependencies, e.g., pandas, sqlite3]

Example:
    ido.add_or_update_item(item_id, item_info)
    stock = ido.get_item_stock(item_id)
    print(f"Current stock: {stock}")

Description:
    - Adds or updates item information in the inventory index during check-in.
    - Updates stock data in the index during check-out.
    - Provides interfaces for querying current stock and item attributes.

This module is designed for the data operation layer of a canteen item
management system, ensuring the accuracy and consistency of inventory data.
"""

import openpyxl
import os
import sys
import __main__
import datetime

import xlrd



def item_data_operate(model, year, month, day, product_name, unit_name, price, quantity, amount, remark, company_name, sigle_name):
    """
    
    在条目表.xlsx中更新或获取条目信息
    
    Parameters:
        model:入库\出库
        year: 入库出库年份
        month: 该条目入库出库的月份
        day: 该条目入库出库的日期
        product_name: 该条目入库出库的食品名
        unit_name: 该条目入库出库的单位
        price: 该条目入库出库的单价
        quantity: 该条目入库出库的数量
        amount: 该条目入库出库的金额
        remark: 该条目入库出库的备注
        company_name: 该条目入库出库的公司名称
        sigle_name: 该条目入库出库的简称

    Returns:
        dict:{品名:[单位,单价,数量,数额,日期]}

    """
    print(f"\nNotice:条目表 {product_name} 数据信息开始更新！")
    wb = openpyxl.Workbook()
    ws = wb.active

    export_data = {product_name: []}

    "物品索引库文件夹是否存在"
    if not os.path.exists(__main__.ITEM_EXCEL_FOLDER):

        "新建索引数据表"

        # 创建条目表文件夹
        os.mkdir(__main__.ITEM_EXCEL_FOLDER)

        # 写入表头
        ws["A1"] = "单位"
        ws["B1"] = "单价"
        ws["C1"] = "存量"
        ws["D1"] = "存值"
        ws["E1"] = "日期"

        # 存为条目表.xlsx 文件
        wb.save(os.path.join(__main__.ITEM_EXCEL_FOLDER , "条目表.xlsx"))

    "打开此索引库"
    item_excel_list = [item for item in os.listdir(__main__.ITEM_EXCEL_FOLDER) if item == "条目表.xlsx"]
    for item in item_excel_list:
        # 若 __main__.ITEM_EXCEL_FOLDER 下 条目表.xlsx 文件存在
        if item == "条目表.xlsx":
            # 打开条目表.xlsx 文件(Mistake: load_workbook 方法采用绝对路径)
            wb = openpyxl.load_workbook(os.path.join(__main__.ITEM_EXCEL_FOLDER , "条目表.xlsx")) # Mistake：

    "检查索引库中是否有此类"
    try:

        worksheet = wb[str(product_name)]

        ws["A1"] = "单位"
        ws["B1"] = "单价"
        ws["C1"] = "存量"
        ws["D1"] = "存值"
        ws["E1"] = "日期"

        # 从D2开始，D列值为C列值与B列值同行相乘的积
        for row in range(2, ws.max_row + 1):
            ws[f"D{row}"] = float(ws[f"C{row}"].value)*float(ws[f"B{row}"].value)

    except KeyError:

        if model == "入库":

            # 创建新Sheet
            wb.create_sheet(product_name)
            ws = wb[product_name]

            ws["A1"] = "单位"
            ws["B1"] = "单价"
            ws["C1"] = "存量"
            ws["D1"] = "存值"
            ws["E1"] = "日期"

            # 从D2开始，D列值为C列值与B列值同行相乘的积
            for row in range(2, ws.max_row + 1):
                ws[f"D{row}"] = float(ws[f"C{row}"].value)*float(ws[f"B{row}"].value)

        elif model == "出库":
            __main__.SAVE_OK_SIGNAL = False
            print("Error: 条目表中无此条目，请先入库！")

    "打开此Sheet"
    ws = wb[product_name]


    "检查操作类型"
    if model == "入库":
        
        "检测除表头外是否有数据"
        if ws.max_row == 1:
            # 追加写入数据
            ws.append([unit_name, price, quantity, amount, f"{year}-{month}-{day}"])

        else:
            "检查价格列中有无匹配条目价格的行"
            # 筛查 B 列中是否有与 price 等值的行
            for row in range(2, ws.max_row + 1):
                if ws[f"B{row}"].value == price:
                    # 更新存量列(C列)、存值列(D列)、日期列(E列)
                    print(f"Notice: 条目表 {product_name} 页更新前第 {row} 行 单位:{ws[f'A{row}'].value} 单价:{ws[f'B{row}'].value} 存量:{ws[f'C{row}'].value} 存值:{ws[f'D{row}'].value} 日期:{ws[f'E{row}'].value}")
                    ws[f"C{row}"] = float(ws[f"C{row}"].value) + float(quantity)
                    ws[f"D{row}"] = float(ws[f"C{row}"].value)*float(ws[f"B{row}"].value)
                    ws[f"E{row}"] = f"{year}-{month}-{day}"        
                    print(f"Notice: 条目表 {product_name} 页更新后第 {row} 行 单位:{ws[f'A{row}'].value} 单价:{ws[f'B{row}'].value} 存量:{ws[f'C{row}'].value} 存值:{ws[f'D{row}'].value} 日期:{ws[f'E{row}'].value}")

    elif model == "出库":

        "暂存该表中的条目"
        data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

        "检测除表头外是否有数据"
        if ws.max_row == 1:
            __main__.SAVE_OK_SIGNAL = False
            print(f"Error: 条目表中 {product_name} 没有存储数据，请先入库！ ")
            return

        else:
            "遍历行列"
        
            for row in range(2, ws.max_row + 1):  
                "比较该行存量与出库数量关系"
                if float(ws[f"C{row}"].value) >= float(quantity):
                    
                    #将出库清单添加到 export_data
                    export_data[product_name].append([ws[f"A{row}"].value,ws[f"B{row}"].value , quantity , float(quantity) * float(ws[f'B{row}'].value), ws[f"E{row}"].value])

                    # 更新存量列(C列)、存值列(D列)、日期列(E列)|Mistake: 单元格在运算的时候需要先进行强制类型转换
                    print(f"Notice: 条目表 {product_name} 页更新前第 {row} 行 单位:{ws[f'A{row}'].value} 单价:{ws[f'B{row}'].value} 存量:{ws[f'C{row}'].value} 存值:{ws[f'D{row}'].value} 日期:{ws[f'E{row}'].value}")
                    ws[f"C{row}"] = float(ws[f"C{row}"].value) - float(quantity)
                    ws[f"D{row}"] = float(ws[f"C{row}"].value)*float(ws[f"B{row}"].value)
                    ws[f"E{row}"] = f"{year}-{month}-{day}"
                    print(f"Notice: 条目表 {product_name} 页更新后第 {row} 行 单位:{ws[f'A{row}'].value} 单价:{ws[f'B{row}'].value} 存量:{ws[f'C{row}'].value} 存值:{ws[f'D{row}'].value} 日期:{ws[f'E{row}'].value}")
                    
                    quantity = 0

                elif float(ws[f"C{row}"].value) < float(quantity):   
                    
                    # 将出库清单添加到 export_data
                    export_data[product_name].append([ws[f"A{row}"].value,ws[f"B{row}"].value , ws[f"C{row}"].value,float(ws[f"B{row}"].value) * float(ws[f'C{row}'].value), ws[f"E{row}"].value])

                    # 更新存量列(C列)、存值列(D列)、日期列(E列)
                    ws[f"C{row}"] = 0
                    ws[f"D{row}"] = float(ws[f"C{row}"].value)*float(ws[f"B{row}"].value)
                    ws[f"E{row}"] = f"{year}-{month}-{day}"

                    quantity = float(quantity) - float(ws[f"C{row}"].value)

                    continue    
                                
                if quantity == 0:
                    break 
            
            "检查库存数是否够出库数"
            if float(quantity) != 0:
                
                "复原条目"
                for row in range(2, ws.max_row + 1):
                    ws[f"C{row}"] = data_rows[row - 2][2]
                    ws[f"D{row}"] = data_rows[row - 2][3]
                    ws[f"E{row}"] = data_rows[row - 2][4]
                
                __main__.SAVE_OK_SIGNAL = False
                print(f"Error: 条目表中 {product_name} 存量不足，请重新输入！ ")
                return
            

    "将从第二行开始的所有行按照单价值从小到大重新排序"
    # 提取数据（不包括表头）
    data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))
    # 按单价（B列，索引1）排序
    data_rows.sort(key=lambda x: x[1])
    # 清除原有数据（不包括表头）
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    # 重新写入排序后的数据
    for idx, row_data in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=idx, column=col_idx, value=value)

    "保存此Sheet"
    wb.save(__main__.ITEM_EXCEL_FOLDER + "条目表.xlsx")
    print(f"Notice: 条目表 {product_name} 页更新完成！")
    return export_data

    
def reindex_item_data():
    
    """
    提取子表主食表,子表副食表中的信息，重新生成条目表.xlsx
    
    Parameters:
        None
        
    Returns:
        None
    """

    year=str(datetime.datetime.now().year)

    print(f"\nNotice: 条目表 数据信息开始重新生成！")

    "删除\work\条目表\条目表.xlsx"
    if os.path.exists(os.path.join(__main__.ITEM_EXCEL_FOLDER , "条目表.xlsx")):
        os.remove(os.path.join(__main__.ITEM_EXCEL_FOLDER , "条目表.xlsx"))
        print(f"Notice: 条目表 数据信息删除完成！")

    "新建work\条目表\条目表.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    "遍历工作簿"
    xls_files = [os.path.join(__main__.SUB_WORK_EXCEL_FOLDER, f) for f in os.listdir(__main__.SUB_WORK_EXCEL_FOLDER) if f.endswith('.xls')]
    
    for file in xls_files:
        
        try:
            workbook = xlrd.open_workbook(file)
            sheet_names = workbook.sheet_names()

            for sheet_name in sheet_names:
                
                worksheet = workbook.sheet_by_name(sheet_name)
                price_group = []                                # 记录连续价格组的三维数组

                "从头到尾记录有单价的行"
                for row_idx in range(1, worksheet.nrows):           # 跳过表头，从第二行开始
                    
                    row = worksheet.row(row_idx)
                    try:       
                        
                        price = float(row[4].value)                                 # 单价
                        month = int(row[0].value) if row[0].value != '' else 0      # 月份
                        day = int(row[1].value) if row[1].value != '' else 0        # 日期
                        price_group.append([row_idx + 1, price,month,day])          # 记录该 [行号,单价,月份,日期]
                        print(f"Notice:  {worksheet.name} 工作簿第 {row_idx + 1} 行 单价列值为{price} ")

                    except Exception as e:
                        print(f"Error: 处理文件 {file} 的 {worksheet.name} 工作簿第 {row_idx + 1} 行时出错: {e}")
                        continue
                
                print(f"Notice:  {worksheet.name} 工作簿价格行信息提取完成，信息为 {price_group}")

                "根据是否为连续行对工作簿价格行进行重新分组"
                if price_group:
                    grouped_prices = []
                    current_group = [price_group[0]]

                    for i in range(1, len(price_group)):
                        if price_group[i][0] == price_group[i-1][0] + 1:  # 检查是否为连续行
                            current_group.append(price_group[i])
                        else:
                            grouped_prices.append(current_group)
                            current_group = [price_group[i]]

                    grouped_prices.append(current_group)  # 添加最后一组
                    print(f"Notice:  {worksheet.name} 工作簿价格行信息重新分组完成，信息为 {grouped_prices}")

                    "遍历分组，根据每一组的最后一个价位对应的条目检查其该行数量列是否存在数量"
                    for group in grouped_prices:
                        
                        last_row_idx = group[-1][0]  # 获取该组最后一个价位的行号 
                        last_price = group[-1][1]
                        quantity_cell = worksheet.cell(last_row_idx - 1, 9)  # 数量列索引为3

                        product_name = ""
                        last_price = 0
                        unit_name = ""

                        try:
                            quantity = float(quantity_cell.value)
                            if quantity > 0:
                                product_name = sheet_name                              # 品名为工作簿名称    
                                last_price =   group[-1][1]                            # 单价为该组最后一个价位 
                                month = group[-1][2]                                   # 月份
                                day = group[-1][3]                                     # 日期                
                                print(f"Notice:  {worksheet.name} 工作簿第 {last_row_idx} 行 有效数量 {quantity}，准备更新条目表，信息为 品名:{product_name} 单位:{unit_name} 单价:{last_price} 数量:{quantity}")
                                
                            else:
                                product_name = sheet_name
                                last_price =   group[-1][1]
                                month = group[-1][2]                                   
                                day = group[-1][3]                                     
                                quantity = 0
                                print(f"Warning: {worksheet.name} 工作簿第 {last_row_idx} 行 数量列值为零或负数，跳过该行。")
                            
                        except Exception as e:
                            print(f"Warning: 处理文件 {file} 的 {worksheet.name} 工作簿第 {last_row_idx} 行时报警，信息为: {e}")
                            continue

                        "为单条目不同价位进行入库操作"
                        try:
                            item_data_operate("入库",year , month, day, product_name, unit_name, last_price, quantity, float(last_price)*float(quantity), f"由 {worksheet.name} 工作簿第 {last_row_idx} 行 入库", "司", "")
                        except Exception as e:
                            print(f"Error: 调用 item_data_operate 方法时出错: {e}")
                            continue
                else:
                    print(f"Warning: {worksheet.name} 工作簿中未找到任何价格信息，跳过该工作簿。")
        
        except Exception as e:
            print(f"Error: {file} 打开失败: {e}")
            continue





