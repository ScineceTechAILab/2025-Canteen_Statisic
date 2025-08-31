import os
import threading
  
import Levenshtein
import xlrd
import pandas as pd

from openpyxl import load_workbook
import paddleocr
from paddleocr import PPStructure
def image_to_excel(
    image_path: str,
    save_folder: str = "./src/data/input/manual",
    ocr_model_path: str = None
    
):
    """
    使用PaddleOCR和PPStructure识别图片中的表格并导出为Excel文件（支持追加写入）。

    :param image_path: 输入图片路径
    :param save_folder: Excel文件保存目录
    :param ocr_model_path: OCR模型路径
    """
    print(f"Notice: 线程 {threading.current_thread().name}(ID={threading.get_ident()})开始处理 {image_path} 的表格识别。")
    os.makedirs(save_folder, exist_ok=True)
    excel_path = os.path.join(save_folder, f"temp_img_input.xlsx")

    # 初始化表格结构识别引擎
    try:
        # 保存原始BASE_DIR值
        original_base_dir = None
        
        # 如果提供了OCR模型路径，则设置PaddleOCR的BASE_DIR
        if ocr_model_path:    
            original_base_dir = paddleocr.BASE_DIR
            paddleocr.BASE_DIR = os.path.abspath(ocr_model_path)
            
        # 恢复原始BASE_DIR值
        else:
            paddleocr.BASE_DIR = original_base_dir
        
        table_engine = PPStructure()
        
    except Exception as e:
        print(f"Error: 线程 {threading.current_thread().name}（ID={threading.get_ident()}）无法初始化表格结构识别引擎,错误信息: {e}")
        return

    # 进行表格结构识别
    result = table_engine(image_path)

    # 提取表格内容
    new_tables = []
    for i, item in enumerate(result):
        if item['type'] == 'table':
            table_html = item['res']['html']
            dfs = pd.read_html(table_html)
            for df in dfs:
                new_tables.append(df)

    # 追加写入逻辑
    if os.path.exists(excel_path):
        # 读取已有数据
        try:
            existing_df = pd.read_excel(excel_path)
        except Exception:
            existing_df = pd.DataFrame()
        # 合并所有新表格
        combined_new = pd.concat(new_tables, ignore_index=True) if new_tables else pd.DataFrame()
        # 合并已有和新数据
        final_df = pd.concat([existing_df, combined_new], ignore_index=True)
    else:
        # 仅保存新表格
        final_df = pd.concat(new_tables, ignore_index=True) if new_tables else pd.DataFrame()

    # 保存到Excel
    final_df.to_excel(excel_path, index=False)
    print(f"Notice: 线程 {threading.current_thread().name}(ID={threading.get_ident()})已完成 {image_path} 的表格识别并追加导出到Excel文件。")

    # 修正图像识别结果表格数据
    print("Notice: 开始修正图像识别结果表格数据")
    
    xls_files = [os.path.join(".", "src", "data", "storage", "work", "子表", f) for f in os.listdir(os.path.join(".", "src", "data", "storage", "work", "子表")) if f.endswith('.xls')]
    res = []
    for file in xls_files:
        try:
            workbook = xlrd.open_workbook(file)
            sheet_names = workbook.sheet_names()
            res += sheet_names
            # print(f"{file} 的 sheets: {sheet_names}")
        except Exception as e:
            print(f"{file} 打开失败: {e}")
    
    products = res
    
    # 打开 Excel 文件
    save_folder = "./src/data/input/manual"
    excel_path = os.path.join(save_folder, f"temp_img_input.xlsx")
    wb = load_workbook(excel_path)
    ws = wb.active# 默认sheet
    
    # 遍历第 C 列（即第 3 列）从第 2 行开始
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        cell = row[0]
        if cell.value is not None and str(cell.value).strip() != "":
            image_product = cell.value
            ratios = dict()  # 每个标准名对应一个近似度，取近似度最大的那个名字
            for product_name in products:
                ratios[product_name] = Levenshtein.ratio(image_product, product_name)
            # 取相似度最大的词
            best_name = max(ratios, key=ratios.get)
            print(f"{cell.value} -- > {best_name}")
            ws.cell(row=cell.row, column=2, value=best_name)

    # 保存修正后的 Excel 文件
    print(f"Notice: 修正后的数据已保存到 {excel_path}")
    wb.save(excel_path)
    wb.close()
    print(f"Notice: 线程 {threading.current_thread().name}(ID={threading.get_ident()})已经完成 {image_path} 的表格识别。")
