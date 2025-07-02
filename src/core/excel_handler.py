# -*- coding: utf-8 -*-
# @Time    : 2025/4/26 19:34
# @Author  : ESJIAN
# @Email   : esjian@outlook.com
# @File    : excel_handler.py
# @Software: VsCode



from glob import glob 
import shutil
import time
from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt, QEvent)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, QFontDatabase, QGradient, QIcon,
    QImage, QKeySequence, QLinearGradient, QPainter,
    QPalette, QPixmap, QRadialGradient, QTransform, Qt)
from PySide6.QtWidgets import (QMessageBox,QAbstractScrollArea,QApplication, QButtonGroup, QFormLayout, QGridLayout,
    QGroupBox, QHBoxLayout, QLabel, QLayout,
    QLineEdit, QPlainTextEdit, QPushButton, QScrollArea,
    QSizePolicy, QSpinBox, QTabWidget, QVBoxLayout,
    QWidget, QFileDialog, QDialog, QVBoxLayout, QCheckBox)


from openpyxl import load_workbook
from datetime import datetime
import os
import subprocess
import sys
import __main__
import openpyxl
import xlrd
from xlutils.copy import copy
from xlwt.Style import  XFStyle


from xlwt import Workbook
import xlwings as xw
import re
import __main__

from src.gui.data_save_dialog import data_save_success
from src.core.excel_handler_utils import (
    is_single_punctuation,
    is_visually_empty,
    is_previous_rows_after_page_break,
    convert_number_to_chinese,
    find_matching_today_rows,
    find_matching_month_rows,
    find_the_first_empty_line_in_main_excel,
    find_the_first_empty_line_in_sub_main_excel,
    find_the_first_empty_line_in_sub_auxiliary_excel,
    get_all_sheets_todo_for_main_table,
    get_all_sheets_todo_for_sub_table
)



from src.core.models.page_counter import *

def store_single_entry_to_temple_excel(self,data, file_path):
    """
    将单条目的数据追加存储到临时excel表格中
    :param data: 要存储的字典数据
    :param file_path: 存储的文件路径
    :return: None
    """
    print("Notice: 将单条目的数据追加存储到临时excel表格中的路径是", file_path)

    if not isinstance(data, dict):
        raise ValueError("数据必须是字典类型")
    
    # 确保字典的值是列表
    for key, value in data.items():
        if not isinstance(value, list):
            data[key] = [value]  # 转换为列表

    # 将字典数据转换为二维列表
    headers = list(data.keys())      # 获取传入字典的键列
    rows = list(zip(*data.values())) # 将键列解包重新打包成每一行的数据

    try:
        if os.path.exists(file_path):
            # 打开现有文件
            workbook = xlrd.open_workbook(file_path, formatting_info=True)
            sheet = workbook.sheet_by_index(0)
            # 过滤掉没有内容的行
            existing_rows = sum(1 for row_idx in range(sheet.nrows) if any(sheet.row_values(row_idx))) # Learning3:条件Sum函数的高级用法

            # 创建可写副本
            writable_workbook = copy(workbook)
            writable_sheet = writable_workbook.get_sheet(0)
            
            # 检查表头是否依次为"日期"、"类别"、"品名"、"单位"、"单价"、"数量"、"金额"、"备注"、"公司"、"单名"
            if sheet.row_values(0) != headers:
                raise ValueError("Error:表头必须为'日期'、'类别'、'品名'、'单位'、'单价'、'数量'、'金额'、'备注'、'公司'、'单名'")

            # 追加数据
            for row_index, row_data in enumerate(rows, start=existing_rows):
                for col_index, cell_value in enumerate(row_data):
                    writable_sheet.write(row_index, col_index, str(cell_value))  # 转为文本存储

            # 保存文件
            writable_workbook.save(file_path)
        else:
            # 创建工作簿对象
            workbook = Workbook()
            # 创建工作表对象
            sheet = workbook.add_sheet("Sheet1")
            
            # 写入表头行数据
            for col_index, header in enumerate(headers):
                sheet.write(0, col_index, header)
            # 写入内容行数据
            for row_index, row_data in enumerate(rows, start=1):
                for col_index, cell_value in enumerate(row_data):
                    sheet.write(row_index, col_index, str(cell_value))  # 转为文本存储

            # 保存文件
            workbook.save(file_path)
            # 打印信息
            print(f"Warning:暂存表格不存在,重新创建暂存表格,路径为{file_path}")

        print("Notice:数据已成功追加存储到Excel文件中。")
        # 显示保存成功的消息提示弹窗
        data_save_success(self)     
    except Exception as e: # Leraning2：不能操作Excel正在打开的表
        print(f"Error:写入Excel文件时出错: {e}")

def clear_temp_xls_excel(self, quit_flag= False):
    """
    清空暂存的 Excel 表格内容
    
    :param: None
    :return: None
    """
    try:
        for i in [__main__.TEMP_SINGLE_STORAGE_EXCEL_PATH, __main__.PHOTO_TEMP_SINGLE_STORAGE_EXCEL_PATH2]:
            print("Notice:正在清空" + i)
            if os.path.exists(i) :
                # 打开现有文件
                workbook = xlrd.open_workbook(i, formatting_info=True)
                writable_workbook = copy(workbook)
                sheet = writable_workbook.get_sheet(0)
                original_sheet = workbook.sheet_by_index(0)

                # 清空内容，只保留表头
                for row_index in range(1, original_sheet.nrows):
                    for col_index in range(original_sheet.ncols):
                        sheet.write(row_index, col_index, "")  # 清空单元格内容
                # 保存文件
                writable_workbook.save(i)
        
        # 触发清空事件若非程序关闭事件，则调整界面显示信息
        if quit_flag == False: 
            # 重新设定暂存项为 0 项
            self.storageNum.setText(QCoreApplication.translate("Form","0", None))
            # 设定正在编辑项为 1 项
            self.spinBox.setValue(1)
            # 复位 TEMP_STORAGED_NUMBER_LISTS
            __main__.TEMP_STORAGED_NUMBER_LISTS = 1
    except Exception as e:
        print(f"Error: 清空暂存表格时出错: {e}")

def clear_temp_xlxs_excel():
    """
    清空暂存的 Excel 表格内容
    :param: None
    :return: None
    """

    # 将目标表格数据全部删除
    temp_xlxs_excel_path = "./src/data/input/manual/temp_img_input.xlsx"
    if os.path.exists(temp_xlxs_excel_path):
        # 打开工作簿
        workbook = load_workbook(temp_xlxs_excel_path)
        # 获取第一个工作表
        sheet = workbook.active
        # 清空工作表内容
        sheet.delete_rows(1, sheet.max_row)
        # 保存工作簿
        workbook.save(temp_xlxs_excel_path)

def clear_temp_image_dir():
    """
    清空图片导入临时目录
    """
    dest_dir = __main__.TEMP_IMAGE_DIR
    files = []
    for _, __, _files in os.walk(dest_dir):
        for file in _files:
            files.append(os.path.join(dest_dir, file))  # 记录完整路径
    for i in files:
        os.remove(i)
    print("Notice:删除临时图片目录成功")


def commit_data_to_storage_excel(self,modle,main_excel_file_path,sub_main_food_excel_file_path,sub_auxiliary_food_excel_file_path,welfare_food_excel_file_path):
    """
    提交暂存 Excel 数据到主表、副表 Excel 文件

    Parameters:
        - self: 当前窗口对象
        - modle: 输入输出模式切换变量,值为 manual/photo
        - main_excel_file_path: 主表 Excel 文件路径
        - sub_main_food_excel_file_path: 主副食明细账 Excel 文件路径
        - sub_auxiliary_food_excel_file_path: 副食品明细账 Excel 文件路径
        - welfare_food_excel_file_path: 福利食品明细账 Excel 文件路径
    :return: None
    """
    "根据触发该函数的是手动还是照片输入模式去读取不同的表格"

    # 调用xlwings库将 xlsx 文件转换为 xls 格式（无头模式）
    with xw.App(visible=False) as app:
        if modle == "manual":
            # 读取手动模式暂存表格
            try:
                # 读取暂存工作簿
                read_temp_storage_workbook = xlrd.open_workbook(__main__.TEMP_SINGLE_STORAGE_EXCEL_PATH)
            except Exception as e:
                print(f"Error: 打开暂存表工作簿出错 {e}")
                return
            
        elif modle == "photo":
            # 读取照片模式暂存表格
            try:
                
                book = app.books.open(__main__.PHOTO_TEMP_SINGLE_STORAGE_EXCEL_PATH)
                book.save(__main__.PHOTO_TEMP_SINGLE_STORAGE_EXCEL_PATH.replace('.xlsx', '.xls'))
                book.close()
                # 读取暂存工作簿
                read_temp_storage_workbook = xlrd.open_workbook(__main__.PHOTO_TEMP_SINGLE_STORAGE_EXCEL_PATH2)
            
            except Exception as e:
                print(f"Error: 打开暂存表工作簿出错 {e}")
                return

        # 获取暂存输入表表头数据
        try:
            read_temp_storage_workbook_headers = read_temp_storage_workbook.sheet_by_index(0).row_values(0)  # 获取表头的第一行数据
            # 确保表头是一个扁平的列表
            if not all(isinstance(header, str) for header in read_temp_storage_workbook_headers):
                raise ValueError("Error: 暂存表头必须是字符串类型")
        except Exception as e:
            print(f"Error: 获取暂存表表头出错,可能 {__main__.TEMP_SINGLE_STORAGE_EXCEL_PATH} 表格为空 {e}")
            return

  
        if __main__.ONLY_WELFARE_TABLE == False:
            "更新主表、子表信息"
            try:
                # 在主表中更新信息
                update_main_table(self,app,main_excel_file_path, read_temp_storage_workbook, read_temp_storage_workbook_headers)
                # 在子表中更新信息
                update_sub_tables(self,app,sub_main_food_excel_file_path, sub_auxiliary_food_excel_file_path, read_temp_storage_workbook, read_temp_storage_workbook_headers)
                #等所有表格都更新完了才日计和月计
                add_counter(self,app ,modle,main_excel_file_path, sub_main_food_excel_file_path, sub_auxiliary_food_excel_file_path,welfare_food_excel_file_path)
        
            except Exception as e:
                __main__.SAVE_OK_SIGNAL = False
                print(f"Error: 更新主表、子表信息出错,错误信息为 {e}")

            self.pushButton_5.setText("提交数据")
            self.pushButton_9.setText("提交数据")
        
        
        else:
            "更新福利表"
            try:
                # 在福利表中更新信息
                update_welfare_food_sheet(self,app,welfare_food_excel_file_path,read_temp_storage_workbook,read_temp_storage_workbook_headers)
                #等所有表格都更新完了才日计和月计
                add_counter(self, app,modle,main_excel_file_path, sub_main_food_excel_file_path, sub_auxiliary_food_excel_file_path,welfare_food_excel_file_path)
            except Exception as e:
                __main__.SAVE_OK_SIGNAL = False
                print(f"Error: 福利表文件读取保存工作失败 {e}")
            
            self.pushButton_5.setText("提交数据")
            self.pushButton_9.setText("提交数据")
            
            # 调用弹窗显示保存完成信息，终端同步显示信息
            print(f"Notice: 文件读取保存工作完成")

        if  __main__.SAVE_OK_SIGNAL:

            # 调用弹窗显示保存完成信息，终端同步显示信息
            print(f"Notice: 主子表文件读取保存工作完成")
            self.worker.done.emit("tables_updated")  # 比如写完数据后调用
        
        else:
            print(f"Error: 主子表文件读取保存工作失败，有条目未识别")

            "依次复原 main、work 目录下的文件"
            # 遍历src\data\storage\backup下的文件夹，找到备份时间最新的一份
            backup_dir = os.path.join('src', 'data', 'storage', 'backup')
            backup_folders = [f for f in glob(os.path.join(backup_dir, '*')) if os.path.isdir(f)]
            backup_folders.sort(key=lambda x: os.path.getmtime(x), reverse=True)
            if backup_folders:
                backup_path = backup_folders[0]
                print(f"Notice: 尝试将最新备份备份 {backup_path} 还原至 main 目录")
            # 还原 backup 目录中的文件到 main、work 目录
            try:
                # 检测是否还存在未关闭的 Excel 进程，若存在则强制关闭防止因为文件被占用而无法删除
                if sys.platform == "win32":
                    # Windows 系统
                    subprocess.call(["taskkill", "/F", "/IM", "EXCEL.EXE"])

                time.sleep(1)  # 等待一秒钟，确保 Excel 进程被关闭

                shutil.rmtree("./src/data/storage/main")      # 删除 main 目录
                shutil.copytree(backup_path, "./src/data/storage/main",dirs_exist_ok=True) # 复制备份目录到 main 目录
                print(f"Notice:文件已从{backup_path}还原到 ./src/data/storage/main 目录")
                
                shutil.rmtree("./src/data/storage/work")      # 删除 work 目录
                shutil.copytree(backup_path, "./src/data/storage/work",dirs_exist_ok=True) # 复制备份目录到 work 目录
                print(f"Notice: 文件已从{backup_path}还原到 ./src/data/storage/work 目录")
            
            except Exception as e:
                print(f"Error: 将主表文件复制到 work 目录出错,错误信息为: {e}")            
            # 弹出入库失败通知
            self.worker.done.emit("tables_updated_filed")  


def update_main_table(self,app,excel_file_path, read_temp_storage_workbook, read_temp_storage_workbook_headers):
    """
    处理主工作簿更新相关表格信息
    :param excel_file_path: 主工作簿路径
    :param read_temp_storage_workbook: 暂存表格工作簿对象
    :param read_temp_storage_workbook_headers: 暂存表格表头
    :return: None
    """

    try:

        main_workbook = app.books.open(excel_file_path)
        print(f"Notice: 主工作表加载成功，文件路径: {excel_file_path}")

        # 创建一个字典，用于存储列索引和列名的对应关系
        header_index = {name: idx for idx, name in enumerate(read_temp_storage_workbook_headers)}
        # 轮询读取暂存表格数据行
        for row_index in range(1, read_temp_storage_workbook.sheet_by_index(0).nrows):
            # 读取行数据
            row_data = read_temp_storage_workbook.sheet_by_index(0).row_values(row_index)
            
            
            try:
                # 将日期分解为月和日
                year, month, day = row_data[header_index["日期"]].split("-")
                # 获取行中类别列类型单元中的类别名数据
                category_name = row_data[header_index["类别"]]
                # 获取行中品名列类型单元中的品名名数据
                product_name = row_data[header_index["品名"]]
                # 获取行中单位列类型单元中的单位名数据
                unit_name = row_data[header_index["单位"]]
                # 获取行中单价列类型单元中的单价名数据
                price = row_data[header_index["单价"]]
                # 获取行中数量列类型单元中的数量名数据
                quantity = row_data[header_index["数量"]]
                # 获取行中金额列单元中金额数据
                amount = row_data[header_index["金额"]]
                # 获取行中备注列单元中备注数据
                remark = row_data[header_index["备注"]]
                # 获取行中公司列单元中公司名数据
                company_name = row_data[header_index["公司"]]
                # 获取行中单名称列单元中单名数据
                single_name = row_data[header_index["单名"]]
            
            except Exception as e:
                __main__.SAVE_OK_SIGNAL = False
                print(f"Error: 处理主工作簿，更新相关表格信息时拆解数据出错 {e}")
            
            if not __main__.MODE:
                print(f"\n\nNotice: {product_name}正在入库")
                # 更新指定公司sheet中的金额数据
                update_company_sheet(self,main_workbook,product_name ,company_name, amount) # 只在入库的时候用到
                # 更新入库相关表中的条目信息
                updata_import_sheet(self,main_workbook, product_name,single_name, row_data, header_index, month, day, unit_name) #TODO:优化耗时
                # 更新食品收发库存表中的条目信息
                update_inventory_sheet(self,main_workbook, product_name, unit_name, quantity, price, amount, remark)
                # 更新收发存表皮中的条目信息
                update_receipt_storage_sheet(self,main_workbook, product_name,single_name, category_name, amount)
                # 更新主副食明细账中的条目信息
                update_main_food_detail_sheet(self,main_workbook, product_name,single_name, category_name, amount)
            else:
                print(f"\n\nNotice: 品名{product_name}正在出库")
                # 此函数不带"export"头，没打错
                updata_import_sheet(main_workbook, single_name, row_data, header_index, month, day, unit_name)
                #搞定
                export_update_inventory_sheet(main_workbook, product_name, unit_name, quantity, price, amount, remark)
                #搞定
                export_update_receipt_storage_sheet(main_workbook, single_name, category_name, amount)
                #搞定
                export_update_main_food_detail_sheet(main_workbook, single_name, category_name, amount)
        
        main_workbook.save()
        "如果开启页计功能，先不关闭主表"
        if not (__main__.ADD_DAY_SUMMARY or __main__.ADD_MONTH_SUMMARY or __main__.ADD_PAGE_SUMMARY or __main__.ADD_TOTAL_SUMMARY):    
            # 关闭工作簿
            main_workbook.close()
        
    except Exception as e:
        print(f"Error: 处理主工作簿更新相关表格信息出错,出错信息{e}")



def update_company_sheet(self,main_workbook, product_name ,company_name, amount):
    """
    更新指定公司sheet中的金额数据
    
    Parameters:
      main_workbook: 主工作簿对象
      company_name: 公司名称
      amount: 要增加的金额
    
    Return: 
      None
    """
    print(f"Notice: 正在更新公司 {company_name} Sheet 页以更新其金额数据")
    
    # 查找对应的公司sheet
    try:
        sheet = main_workbook.sheets[company_name]
        
        "更新阿拉伯数字的值"
        # 获取当前值
        current_value = sheet.range("D8").value  # 一般公司 Sheet 金额单元格是 D8（Excel索引从1开始）
        if current_value is None or current_value == "":
            current_value = 0

        # 读取金额
        if amount is None or amount == "":
            amount = 0

        "检查金额格式"
        try:
            # 去掉金额字符中的空字符，防止出现类似 '10. 56' 的字符造成强制类型转换报错
            if isinstance(amount, str):  # 确保 amount 是字符串
                amount = amount.replace(" ", "")  # 去掉字符串首尾的空字符
            
            amount = float(amount)
        except Exception:
            print(f"Error: 公司 Sheet 的传入金额数据格式错误，请检查输入金额是否正确，已跳过公司金额数据的写入")
            amount = 0
            return


        "计算新值"
        try:
            if isinstance(current_value, (int, float)):
                new_value = current_value + amount
            else:
                try:
                    new_value = float(current_value) + amount
                except Exception:
                    new_value = amount
            new_value = round(new_value, 2)  # 保留两位小数
            print("Notice: 公司金额当前值", current_value, "新值", new_value)
        except Exception as e:
            print(f"Error: 更新公司 Sheet 金额数据中计算新值步骤出错 {e}，已跳过公司金额数据的写入")
            
            return

        "写入新值"
        try:
            # 写入新值
            sheet.range("D8").value = new_value
            #更新汉字大写数字的值
            new_value_chinese = convert_number_to_chinese(new_value)  # 转换为中文大写金额
            print("Notice: 当前值", current_value, "新值", new_value_chinese)
            sheet.range("L8").value = new_value_chinese
            print(f"Notice: 在公司名为 {company_name} 的sheet中更新金额数据成功, 新值为 {new_value_chinese}")
        except Exception as e:
            print(f"Error: 更新公司 Sheet 金额数据中写入新值步骤出错 {e}，已跳过公司金额数据的写入")
        
            return
        
    except Exception as e: # KeyError 会造成直接退出该函数步骤
        print(f"Warning: 更新指定公司 {company_name} Sheet 失败，{e}")
        return



def updata_import_sheet(self,main_workbook, product_name,single_name, row_data, header_index, month, day, unit_name):
    """
    将数据写入指定的入库类型sheet中
    
    Parameters:
        :param main_workbook: 主工作簿对象
        :param single_name: sheet名称
        :param row_data: 行数据
        :param header_index: 表头索引字典
        :param month: 月份
        :param day: 日期
        :param unit_name: 单位名称
        :return: None
    
    主表各种杂表
    wjwcj 2025/05/05 13:47 测试没问题
    """
    print(f"Notice: 正在查询入库类型名为 {single_name} 的sheet页以更新其数据")
    try:
        # 检查目标Sheet名是否存在
        if single_name in [s.name for s in main_workbook.sheets]:
            sheet = main_workbook.sheets[single_name]
            print(f"Notice: 在主表中找到入库类型名为 `{single_name}` 的sheet")
        
        elif single_name+" " in [s.name for s in main_workbook.sheets]:
            sheet = main_workbook.sheets[f"{single_name} "]
            print(f"Notice: 在主表中找到入库类型名为 `{single_name} ` 的sheet")
        else:
            print(f"Error: 将数据写入指定的入库类型sheet中时未在主表中找到入库类型名为 `{single_name}` 的sheet")
            #QMessageBox.warning(None, "警告", f"未在主表中找到入库类型名为 `{single_name}` 的sheet,可能存在空字符,已跳过写入指定入库类型 sheet 中步骤")
            return

        

        # 查找第一行空行，记录下空行行标（从表格的第二行开始）
        for row_index in range(0, sheet.used_range.rows.count):
            if sheet.range((row_index + 1, 1)).value is None and row_index != 0:
                # 检查前一行是否包含“领导”二字
                if row_index > 0:
                    previous_row_values = [
                    str(sheet.range((row_index, col)).value).strip()
                    for col in range(1, sheet.used_range.columns.count + 1)
                    if sheet.range((row_index, col)).value is not None
                    ]
                    if any("领导" in value for value in previous_row_values):
                        print(f"Notice: 第 {row_index} 行包含“领导”二字，继续查找下一行")
                        continue

                # 检查当前列的前几行是否包含“序号”二字
                column_values = [
                    str(sheet.range((row, 1)).value).strip()
                    for row in range(1, row_index + 1)
                    if sheet.range((row, 1)).value is not None
                ]
                if not any("序号" in value for value in column_values):
                    print(f"Notice: 前 {row_index} 行未找到“序号”二字，继续查找下一行")
                    continue
                break

        # 尝试写入一行数据
        try:
            # 获取当前列中所有的序号值，排除空值并转换为整数
            existing_numbers = []
            for i in range(row_index):
                box_value = sheet.range((i + 1, 1)).value
                try:
                    # 判断 box_value 是否是"日记"，是则清空已有序号数据
                    if box_value == '日计':
                        existing_numbers = [0]
                    else:
                        box_value = int(box_value)  # 尝试将值转换为整数
                        if str(box_value).isdigit():
                            existing_numbers.append(box_value)  # 如果转换成功，则添加到列表中
                except:
                    continue
            # 计算新的序号值
            new_number = max(existing_numbers) + 1 if existing_numbers else 1
            # 写入序号数据
            sheet.range((row_index + 1, 1)).value = new_number
            print(f"Notice: 在主表为入/出库类型 {single_name} 的第 {row_index} 行写入序号：{new_number} 成功")

            # 为B、C列写入月份日期数据
            sheet.range((row_index + 1, 2)).value = month
            sheet.range((row_index + 1, 3)).value = day
            print(f"Notice: 在主表为入/出库类型 {single_name} 的第 {row_index} 行写入月份：{month} 日：{day} 成功")

            #动态获取表头行行数
            name_row = 0
            for i in range(6):
                #在前六行里找吧
                datas = [str(sheet.range((i + 1, col)).value).strip().replace(" ", "") for col in range(1, 12)]
                if "单价" in datas and "数量" in datas and "金额" in datas:
                    name_row = i + 1
                    break

            # 依次为D~K列写入数据(D、E列合并，需要加入跳过判断逻辑)
            for col_index in range(4, 12):
                if col_index == 5:
                    # 如果当前列是E列，则跳过
                    continue
                else:
                    # 操作该单元时候，访问第该单元对应列的第四行单元获取该列的列名属性
                    # wjwcj：这可不一定，自购主食出库的列名属性就在第三行，所以得动态获取(到name_row)
                        
                    cell_attribute = sheet.range((name_row, col_index)).value

                    if isinstance(cell_attribute, str):
                        # 去除所有中文之间的空格
                        cell_attribute = re.sub(r'(?<=[\u4e00-\u9fa5])\s+(?=[\u4e00-\u9fa5])', '', cell_attribute)

                    try:
                        if cell_attribute == "计量单位":
                            # 如果该列名是单独的计量单位，手动匹配暂存表格中名为单位列的对应单元值
                            sheet.range((row_index + 1, col_index)).value = unit_name
                            print(f"Notice: 在主表为入/出库类型 {single_name} 的 {row_index} 行名为 {cell_attribute} 的列写入值 {row_data[header_index['单位']]} 成功")

                        else:
                            # 在row_data中查找该列名对应的值，然后写入正在被操作的该单元中
                            #print("正在写入" + cell_attribute + "  " + str(row_data[header_index[cell_attribute]]))
                            if cell_attribute == "类别" and single_name  in ["自购主食入库等", "自购主食出库"]:
                                row_data[header_index[cell_attribute]] = row_data[header_index[cell_attribute]] + single_name.strip("等").strip("自购主食")
                            sheet.range((row_index + 1, col_index)).value = row_data[header_index[cell_attribute]]
                            print(f"Notice: 在主表为入/出库类型 {single_name} 的 {row_index} 行名为 {cell_attribute} 的列写入值 {row_data[header_index[cell_attribute]]} 成功")

                    except KeyError:
                        print(f"Error: 未在主表入/出库类型 {single_name} 找到名为 {cell_attribute} 的列")

        except Exception as e:
            print(f"Error: 写入主表时出错 {e}")

    except Exception:
        print(f"Warning: 将数据写入指定的入库类型 `{single_name}` sheet发生错误,已跳过将数据写入指定的入库类型sheet中")



def update_inventory_sheet(self,main_workbook, product_name, unit_name, quantity, price, amount, remark):
    """
    更新或添加数据到食堂物品收发存库存表(入库)
    :param main_workbook: 主工作簿对象
    :param product_name: 物品名称
    :param unit_name: 计量单位
    :param quantity: 数量
    :param price: 单价
    :param amount: 金额
    :param remark: 备注
    :return: None
    """
    # 尝试打开名为食堂物品收发库存表的 sheet
    if "食堂物品收发存库存表" in [s.name for s in main_workbook.sheets]:
        sheet = main_workbook.sheets["食堂物品收发存库存表"]
        print(f"Notice: 找到出库类型名为 `食堂物品收发存库存表` 的sheet")
    else:
        print(f"Error: 未找到出库类型名为 `食堂物品收发存库存表` 的sheet,可能存在空字符")
        # 弹窗提示条目入库出现问题，该条目入库跳过更新更新或添加数据到食堂物品收发存库存表这一步骤
        return
        

    try:
        # 调用Excel VBA API 查找名为'名称'的 A列中是否存在该名称
        found = sheet.range("A:A").api.Find(product_name)

        if found is not None:
            # 如果存在，则更新该行的数据
            # 用遍历方式查找行索引，避免直接用 .row
            row_index = None
            for i in range(1, sheet.used_range.rows.count + 1):
                if sheet.range(f"A{i}").value == product_name:
                    row_index = i
                    break
            if row_index is None:
                print(f"Error: 在表 食堂物品收发存库存表 未找到名称为 {product_name} 的行")
                return
            print(f"Notice: 在表 食堂物品收发存库存表 找到名称为 {product_name} 的行,行号为{row_index}")

            # 判断quantity、price、amount的值是否为数值
            try:
                # 去掉金额字符中的空字符，防止出现类似 '10. 56' 的字符造成强制类型转换报错
                if isinstance(amount, str):
                    amount = amount.replace(" ", "")
                if isinstance(quantity, str):
                    quantity = quantity.replace(" ", "")
                if isinstance(price, str):
                    price = price.replace(" ", "")

                quantity = float(quantity)
                price = float(price)
                amount = float(amount)
            except:
                print(f"Error: quantity、price、amount的值必须为数值")
                return
            
            if isinstance(quantity, (int, float)) and isinstance(price, (int, float)) and isinstance(amount, (int, float)):
                # 在F列更新数量信息，G列更新单价信息，H列更新金额信息
                raw_value = []

                for alpha in "FGH":
                    if sheet.range(f"{alpha}{row_index}").value is None:
                        raw_value.append(0)
                    elif "." in str(sheet.range(f"{alpha}{row_index}").value):
                        raw_value.append(float(sheet.range(f"{alpha}{row_index}").value))
                    else:
                        raw_value.append(int(sheet.range(f"{alpha}{row_index}").value))

                sheet.range(f"F{row_index}").value = raw_value[0] + quantity
                sheet.range(f"G{row_index}").value = raw_value[1] + price
                sheet.range(f"H{row_index}").value = raw_value[2] + amount

                print(f"Notice: 在表 食堂物品收发存库存表 更新行信息 数量、单价、金额 的列,行号为{row_index}")
            else:

                print(f"Error: quantity、price、amount的值必须为数值")
                return
        else:
            # 如果不存在，查找第一行空行，记录下空行
            for row_index in range(0, sheet.used_range.rows.count):
                if sheet.range((row_index + 1, 1)).value is None and sheet.range((row_index + 1, 2)).value is None and sheet.range((row_index + 2, 1)).value is None and sheet.range((row_index + 2, 2)).value is None:
                    break

            # 更新该行A列的物品名称信息
            sheet.range((row_index + 1, 1)).value = product_name
            # 更新该行B列物品的计量单位信息
            sheet.range((row_index + 1, 2)).value = unit_name
            # 更新该行O列物品的备注信息
            sheet.range((row_index + 1, 15)).value = remark

            try:
                # 将quantity、price、amount转换为浮点数
                # 去掉金额字符中的空字符，防止出现类似 '10. 56' 的字符造成强制类型转换报错
                if isinstance(amount, str):
                    amount = amount.replace(" ", "")
                if isinstance(quantity, str):
                    quantity = quantity.replace(" ", "")
                if isinstance(price, str):
                    price = price.replace(" ", "")

                quantity = float(quantity)
                price = float(price)
                amount = float(amount)
            except ValueError:

                print(f"Error: quantity、price、amount的值必须为数值")
                return

            # 在F列更新数量信息，G列更新单价信息，H列更新金额信息
            sheet.range((row_index + 1, 6)).value = quantity
            sheet.range((row_index + 1, 7)).value = price
            sheet.range((row_index + 1, 8)).value = amount

            print(f"Notice: 在表 食堂物品收发存库存表 为 `名称{product_name}、数量 {quantity}、单价 {price}、金额 {amount}` 列添加值,行号为{row_index+1}")

    except Exception as e:
        print(f"Error: 更新食堂物品收发存库存表时出错 {e}")



def update_receipt_storage_sheet(self,main_workbook, product_name,single_name, category_name, amount):
    """
    更新收发存表皮中的条目信息
    :param main_workbook: 主工作簿对象
    :param single_name: 单名信息
    :param category_name: 类别信息
    :param amount: 金额数据
    :return: None
    """
    print(f"Notice: 正在更新主表收发存表皮 Sheet  信息,该条目单名为 '{single_name}'")
    # 尝试打开名为收发存表皮的 sheet
    if "收发存表皮" in [s.name for s in main_workbook.sheets]:
        sheet = main_workbook.sheets["收发存表皮"]
        print(f"Notice: 找到入库类型名为 `收发存表皮` 的sheet")
    else:
        __main__.SAVE_OK_SIGNAL = False
        print(f"Error: 更新收发存表皮表数据时，未在主表找到名为 `收发存表皮` 的sheet,可能存在空字符")
        return

    # 提取输入数据的单名信息和类别信息进行行索引词匹配
    row_index_name = None
    if single_name == "扶贫主食入库":
        row_index_name = "主食（帮扶食品）"

    elif single_name == "扶贫副食入库":
        row_index_name = "副食（帮扶食品）"
    elif single_name in ["自购主食入库", "自购主食入库等","食堂副食入库","食堂主食入库",]:
        if "主食" in category_name:
            row_index_name = "主食（自购）"
        elif "副食" in category_name:
            row_index_name = "副食（自购）"
        else:
            print("Error: 自购主食入库 未找到类别信息，请检查输入数据")
    
    elif single_name == "场调面食入库":
        if "主食" in category_name:
            row_index_name = "正常厂主食"
        elif "副食" in category_name:
            row_index_name = "正常厂副食"
        else:
            print("\nError: 场调面食入库 未找到类别信息，请检查输入数据\n")
    else:
        __main__.SAVE_OK_SIGNAL = False
        print(f"Error: 在更新收发存表皮时规则未匹配到表名 `{single_name}` ,可能存在空字符,已跳过写入")
        return

    # 调用Excel API用行索引名匹配行索引
    found_row = sheet.range("A:A").api.Find(row_index_name)
    if found_row is not None:
        try:
            found_row_index = sheet.range("A:A").value.index(row_index_name) + 1
            print(f"Notice: 在 收发存表皮 Sheet 中找到 {row_index_name} 的行索引为 {found_row_index}")
        except Exception as e:
            print(f"Error: 获取行索引出错 {e}")
            return
    else:
        print(f"Error: 在 收发存表皮 Sheet 中未找到 {row_index_name} 的行索引，请检查输入数据")
        return

    # 更新H列的金额数据
    if found_row_index:
        if sheet.range(found_row_index, 8).value is None:
            sheet.range(found_row_index, 8).value = float(amount)
            print(f"Notice: 发现 收发存表皮 Sheet   中 {row_index_name} 的金额数据不存在,现在添加数据为 {sheet.range(found_row_index, 8).value}")
        else:
            print(f"Notice: 在 收发存表皮 Sheet 中 {row_index_name} 的金额原始数据为 {sheet.range(found_row_index, 8).value}")
            sheet.range(found_row_index, 8).value = float(amount) + float(sheet.range(found_row_index, 8).value) # Fixed：修复了can only concatenate str (not "float") to str，遇到运算问题时尽可能的强制类型转换
            print(f"Notice: 在 收发存表皮 Sheet 中更新 {row_index_name} 的金额数据成功,现在数据为 {sheet.range(found_row_index, 8).value}")
    else:
        print(f"Error: 在 收发存表皮 Sheet 中更新 {row_index_name} 的金额数据失败，请检查输入数据")



def update_main_food_detail_sheet(self,main_workbook,product_name ,single_name, category_name, amount):
    """
    更新主表中主副食品明细账Sheet中的信息
    :param main_workbook: 主工作簿对象
    :param single_name: 单名信息
    :param category_name: 类别信息
    :param amount: 金额数据
    :return: None
    """

    """注意！！这个函数只负责主副食品明细账"""

    try:
        # 尝试打开名为主副食品明细账的 sheet
        if "主副食品明细账" in [s.name for s in main_workbook.sheets]:
            sheet = main_workbook.sheets["主副食品明细账"]
            print(f"Notice: 找到入库类型名为 `主副食品明细账` 的sheet")
        else:
            __main__.SAVE_OK_SIGNAL = False
            print(f"Error: 更新主副食品明细账时未在主表找名为 `主副食品明细账` 的sheet,可能存在空字符")
            return
    except Exception as e:
        print(f"Error: 打开主副食品明细账 Sheet 时出错 {e}")
        return

    # 提取输入数据的单名信息和类别信息进行行列索引词匹配（）
    if single_name == "扶贫主食入库":
        row_index_name = "（帮扶食品）主副食"
        column_index_name = "主食购入"

    elif single_name in "自购主食入库等":
        if category_name == "主食":
            row_index_name = "自购主副食"
            column_index_name = "主食购入"
        elif category_name == "副食":
            row_index_name = "自购主副食"
            column_index_name = "副食购入"
        else:
            print(f"Error: 查找 '自购主食入库' sheet时未找到对应的类别信息，请检查类别")
            return
    
    elif single_name == "扶贫副食入库":
        row_index_name = "（帮扶食品）主副食"
        column_index_name = "副食购入"
    
    elif single_name == "食堂副食入库":
        row_index_name = "（帮扶食品）主副食"
        column_index_name = "副食购入"

    else:
        __main__.SAVE_OK_SIGNAL = False
        print(f"Error: 更新主副食品明细账时未在主表找名为 `{single_name}` 的sheet,可能存在空字符")
        return

    # 调用Excel API 进行行索引名匹配
    found_row = sheet.range("A:A").api.Find(row_index_name)
    if found_row is not None:
        try:
            found_row_index = sheet.range("A:A").value.index(row_index_name) + 1
            print(f"Notice: 在 主副食品明细账 Sheet中找到 {row_index_name} 的行索引为 {found_row_index}")
        except Exception as e:
            print(f"Error: 获取行索引出错 {e}")
            return
    else:
        print(f"Error: 在 主副食品明细账 Sheet中未找到 {row_index_name} 的行索引，请检查输入数据")
        return

    # 调用Excel API 进行列索引名匹配
    found_column = sheet.range("5:5").api.Find(column_index_name)
    if found_column is not None:
        try:
            found_column_index = found_column.Column
            print(f"Notice: 在 主副食品明细账 Sheet中找到 {column_index_name} 的列索引为 {found_column_index}")
        except Exception as e:
            print(f"Error: 获取列索引出错 {e}")
            return
    else:
        print(f"Error: 在 主副食品明细账 Sheet中未找到 {column_index_name} 的列索引，请检查输入数据")
        return

    # 更新相应单元的金额数据
    if found_column_index is not None and found_row_index is not None:
        cell_value = sheet.range(found_row_index, found_column_index).value
        if cell_value is None:
            sheet.range(found_row_index, found_column_index).value = float(amount)
            print(f"Notice: 在 主副食品明细账 Sheet中的 {row_index_name} {column_index_name} 的金额数据为空，已更新为 {amount}")
        else:
            print(f"Notice: 在 主副食品明细账 Sheet中的 {row_index_name} {column_index_name} 的原始金额数据为 {cell_value}")
            sheet.range(found_row_index, found_column_index).value = float(amount) + float(cell_value)
            print(f"Notice: 在 主副食品明细账 Sheet中的 {row_index_name} {column_index_name} 的现在金额数据为 {sheet.range(found_row_index, found_column_index).value}")



def update_sub_tables(self,app,sub_main_food_excel_file_path, sub_auxiliary_food_excel_file_path, read_temp_storage_workbook, read_temp_storage_workbook_headers):
    """
    同时更新子表主食表和副食表
    :param sub_main_food_excel_file_path: 子表主食表路径
    :param sub_auxiliary_food_excel_file_path: 子表副食表路径
    :param read_temp_storage_workbook: 暂存表格工作簿对象
    :param read_temp_storage_workbook_headers: 暂存表格表头
    :return: None
    """



    try:

        # 打开子表主食表
        main_workbook = app.books.open(sub_main_food_excel_file_path)
        # 打开子表副食表
        auxiliary_workbook = app.books.open(sub_auxiliary_food_excel_file_path)
        print(f"Notice: 子表主食表和副食表加载成功，文件路径: {sub_main_food_excel_file_path} 和 {sub_auxiliary_food_excel_file_path}")
        
        print(f"\n\n\nNotice: 入库更新子表")
        # 在子表主食表中更新信息
        update_sub_main_food_sheet(main_workbook, read_temp_storage_workbook, read_temp_storage_workbook_headers)
        # 在子表副食表中更新信息
        update_sub_auxiliary_food_sheet(auxiliary_workbook, read_temp_storage_workbook, read_temp_storage_workbook_headers)
    
    except Exception as e:

        print(f"Error: 更新子表主食表或副食表出错 {e}")


    # 保存工作簿
    main_workbook.save()
    auxiliary_workbook.save()
    "后续如果有页计功能，先不关闭页计表格"
    if not __main__.ADD_DAY_SUMMARY and not __main__.ADD_MONTH_SUMMARY and not __main__.ADD_PAGE_SUMMARY and not __main__.ADD_TOTAL_SUMMARY:
        # 关闭工作簿
        main_workbook.close()
        auxiliary_workbook.close()

    print(f"Notice: 子表主食表和副食表更新完成，已保存并关闭工作簿")


def update_sub_main_food_sheet(main_workbook, read_temp_storage_workbook, read_temp_storage_workbook_headers):
    """
    将暂存表数据提交到子表主食表
    Parameters:
     main_workbook: 子表主食表工作簿对象
     sub_mian_food_excel_file_path: 子表主食表路径
     read_temp_storage_workbook: 暂存表格工作簿对象
     read_temp_storage_workbook_headers: 暂存表格表头
    
    :return: None
    """

    if not __main__.MODE:

        "子表主食表入库模式"
        try:

            # 轮询读取暂存表格数据行
            for row_index in range(1, read_temp_storage_workbook.sheet_by_index(0).nrows):
                # 读取行数据
                row_data = read_temp_storage_workbook.sheet_by_index(0).row_values(row_index)
                # 创建一个字典，用于存储列索引和列名的对应关系
                header_index = {name: idx for idx, name in enumerate(read_temp_storage_workbook_headers)}
                
                try:
                    # 将日期分解为月和日
                    year, month, day = row_data[header_index["日期"]].split("-")
                    # 获取行中类别列类型单元中的类别名数据
                    category_name = row_data[header_index["类别"]]
                    # 获取行中品名列类型单元中的品名名数据
                    product_name = row_data[header_index["品名"]]
                    # 获取行中单位列类型单元中的单位名数据
                    unit_name = row_data[header_index["单位"]]
                    # 获取行中单价列类型单元中的单价名数据
                    price = row_data[header_index["单价"]]
                    # 获取行中数量列类型单元中的数量名数据
                    quantity = row_data[header_index["数量"]]
                    # 获取行中金额列单元中金额数据
                    amount = row_data[header_index["金额"]]
                    # 获取行中备注列单元中备注数据
                    remark = row_data[header_index["备注"]]
                    # 获取行中公司列单元中公司名数据
                    company_name = row_data[header_index["公司"]]
                    # 获取行中单名称列单元中单名数据
                    single_name = row_data[header_index["单名"]]  
                except Exception as e:
                    __main__.SAVE_OK_SIGNAL = False
                    print(f"Error: 将暂存表数据提交到子表主食表时拆解数据出错 {e}")

                if category_name == "主食":
                    # 获取所有sheet的name
                    sheet_names = [s.name for s in main_workbook.sheets]
                    # 筛选包含product_name的sheet名字
                    matching_sheets = [name for name in sheet_names if product_name in re.sub(r'\d+', '', name)]
                    print(matching_sheets)
                    # 取大于product_name长度且长度最小的sheet_name
                    if matching_sheets:
                        sheet_name = min((name for name in matching_sheets if len(re.sub(r'\d+', '', name)) >= len(product_name)), key=len, default=None)
                        if sheet_name:
                            sheet = main_workbook.sheets[sheet_name]
                        else:
                            print(f"未找到合适的sheet匹配品名 {product_name}")
                            return
                    else:
                        print(f"Warning: 未找到品名为 {product_name} 的sheet")
                        return

                    #暂时感觉这个for循环没什么问题
                    #wjwcj: 2025/05/04 15:31
                    for sub_row_index in range(sheet.used_range.rows.count):
                        # 检查每行的1到11列是否都是空
                        if all(is_visually_empty(sheet.range((sub_row_index + 1, col))) for col in range(1, 12)):

                            # 向前检查是否是“过次页 + 空行 + 空行”的模式
                            if is_previous_rows_after_page_break(sheet, sub_row_index + 1):
                                print(f"Warning: 忽略第 {sub_row_index + 1} 行（前面是‘过次页’+连续空行）")
                                continue

                            print("这里开始执行", str(sub_row_index + 1))   

                            # 检查前一行是否符合某些条件（仅包含空格或单个标点符号）
                            if sub_row_index > 0 and all(
                                ((sheet.range((sub_row_index, col)).value is None) or 
                                (is_single_punctuation(str(sheet.range((sub_row_index, col)).value).strip())))
                                for col in range(1, 12)
                            ):
                                print(f"Notice: 发现第 {sub_row_index + 1} 行可用(仅包含空格或单个标点)，开始写入数据")
                                break

                            print(f"Notice: 发现第 {sub_row_index + 1} 行为空行，开始写入数据")
                            break

                    
                    # 往该没有内容的行的A列中写入月份、B列中写入日
                    """
                    !!!注意！！这里原为sub_row_index+1，改为sub_row_index, 因为上文for循环代码认为sub_row_index行已经是可用的了
                    """
                    try:
                        sheet.range((sub_row_index + 1, 1)).value = month
                        sheet.range((sub_row_index + 1, 2)).value = day
                        print(f"Notice: 子表主食表 {product_name} sheet 写入日期成功")
                    except Exception as e:
                        print(f"Error: 子表主食表 {product_name} sheet 写入日期失败{e}")
                        return
                    # 往该没有内容的行的D列中写入出入库摘要
                    try:
                        sheet.range((sub_row_index + 1, 4)).value = "入库"
                        print(f"Notice: 子表主食表 {product_name} sheet 写入出入库摘要成功")
                    except Exception as e:
                        print(f"Error: 子表主食表 {product_name} sheet 写入出入库摘要失败{e}")
                        return
                    # 往该没有内容的行中的E列写入单价
                    try:
                        sheet.range((sub_row_index + 1, 5)).value = price
                        print(f"Notice: 子表主食表 {product_name} sheet 写入单价成功")
                    except Exception as e:
                        print(f"Error: 子表主食表 {product_name} sheet 写入单价失败{e}")
                        return
                    # 往该没有内容的行中的F列写入数量
                    try:
                        sheet.range((sub_row_index + 1, 6)).value = quantity
                        print(f"Notice: 子表主食表 {product_name} sheet 写入数量成功")
                    except Exception as e:
                        print(f"Error: 子表主食表 {product_name} sheet 写入数量失败{e}")
                        return
                    # 往该没有内容的行中的G列写入金额
                    try:
                        sheet.range((sub_row_index + 1, 7)).value = amount
                        print(f"Notice: 子表主食表 {product_name} sheet 写入金额成功")
                    except Exception as e:
                        print(f"Error: 子表主食表 {product_name} sheet 写入金额失败{e}")
                        return
                else:
                    print(f"Warning: 在子表主食表入库时该食品类别属性不名为 主食 ,实名为 {category_name} 已跳过该菜品子表主食表写入")
                    continue
                
        except Exception as e:
            print(f"Error: 将暂存表数据提交到子表主食表时出错,出错信息{e}")
            return  

    else:
        "子表主食表出库模式"
        try:
            # 轮询读取暂存表格数据行
            for row_index in range(1, read_temp_storage_workbook.sheet_by_index(0).nrows):
                # 读取行数据
                row_data = read_temp_storage_workbook.sheet_by_index(0).row_values(row_index)
                # 创建一个字典，用于存储列索引和列名的对应关系
                header_index = {name: idx for idx, name in enumerate(read_temp_storage_workbook_headers)}
                
                try:
                    # 将日期分解为月和日
                    year, month, day = row_data[header_index["日期"]].split("-")
                    # 获取行中类别列类型单元中的类别名数据
                    category_name = row_data[header_index["类别"]]
                    # 获取行中品名列类型单元中的品名名数据
                    product_name = row_data[header_index["品名"]]
                    # 获取行中单位列类型单元中的单位名数据
                    unit_name = row_data[header_index["单位"]]
                    # 获取行中单价列类型单元中的单价名数据
                    price = row_data[header_index["单价"]]
                    # 获取行中数量列类型单元中的数量名数据
                    quantity = row_data[header_index["数量"]]
                    # 获取行中金额列单元中金额数据
                    amount = row_data[header_index["金额"]]
                    # 获取行中备注列单元中备注数据
                    remark = row_data[header_index["备注"]]
                    # 获取行中公司列单元中公司名数据
                    company_name = row_data[header_index["公司"]]
                    # 获取行中单名称列单元中单名数据
                    single_name = row_data[header_index["单名"]]  
                except Exception as e:
                    
                    __main__.SAVE_OK_SIGNAL = False
                    print(f"Error: 将暂存表数据提交到子表主食表(出库)时拆解数据出错 {e}")


                # 获取所有sheet的name
                sheet_names = [s.name for s in main_workbook.sheets]
                # 筛选包含product_name的sheet名字
                matching_sheets = [name for name in sheet_names if product_name in re.sub(r'\d+', '', name)]
                print(matching_sheets)
                # 取大于product_name长度且长度最小的sheet_name
                if matching_sheets:
                    sheet_name = min((name for name in matching_sheets if len(re.sub(r'\d+', '', name)) >= len(product_name)), key=len, default=None)
                    if sheet_name:
                        sheet = main_workbook.sheets[sheet_name]
                    else:
                        print(f"未找到合适的sheet匹配品名 {product_name}")
                        return
                else:
                    print(f"Warning: 未找到品名为 {product_name} 的sheet")
                    return

                #暂时感觉这个for循环没什么问题
                #wjwcj: 2025/05/04 15:31
                for sub_row_index in range(sheet.used_range.rows.count):
                    # 检查每行的1到11列是否都是空
                    if all(is_visually_empty(sheet.range((sub_row_index + 1, col))) for col in range(1, 12)):

                        # 向前检查是否是“过次页 + 空行 + 空行”的模式
                        if is_previous_rows_after_page_break(sheet, sub_row_index + 1):
                            print(f"Warning: 忽略第 {sub_row_index + 1} 行（前面是‘过次页’+连续空行）")
                            continue

                        print("这里开始执行", str(sub_row_index + 1))   

                        # 检查前一行是否符合某些条件（仅包含空格或单个标点符号）
                        if sub_row_index > 0 and all(
                            ((sheet.range((sub_row_index, col)).value is None) or 
                            (is_single_punctuation(str(sheet.range((sub_row_index, col)).value).strip())))
                            for col in range(1, 12)
                        ):
                            print(f"Notice: 发现第 {sub_row_index + 1} 行可用(仅包含空格或单个标点)，开始写入数据")
                            break

                        print(f"Notice: 发现第 {sub_row_index + 1} 行为空行，开始写入数据")
                        break

                
                # 往该没有内容的行的A列中写入月份、B列中写入日
                """
                !!!注意！！这里原为sub_row_index+1，改为sub_row_index, 因为上文for循环代码认为sub_row_index行已经是可用的了
                """
                try:
                    sheet.range((sub_row_index + 1, 1)).value = month
                    sheet.range((sub_row_index + 1, 2)).value = day
                    print(f"Notice: 子表主食表 {product_name} sheet 写入日期成功")
                except Exception as e:
                    print(f"Error: 子表主食表 {product_name} sheet 写入日期失败{e}")
                    return
                # 往该没有内容的行的D列中写入出入库摘要
                try:
                    sheet.range((sub_row_index + 1, 4)).value = "出库"
                    print(f"Notice: 子表主食表 {product_name} sheet 写入出入库摘要成功")
                except Exception as e:
                    print(f"Error: 子表主食表 {product_name} sheet 写入出入库摘要失败{e}")
                    return
                # 往该没有内容的行中的E列写入单价
                try:
                    sheet.range((sub_row_index + 1, 5)).value = price
                    print(f"Notice: 子表主食表 {product_name} sheet 写入单价成功")
                except Exception as e:
                    print(f"Error: 子表主食表 {product_name} sheet 写入单价失败{e}")
                    return
                # 往该没有内容的行中的F列写入数量
                try:
                    sheet.range((sub_row_index + 1, 8)).value = quantity
                    print(f"Notice: 子表主食表 {product_name} sheet 写入数量成功")
                except Exception as e:
                    print(f"Error: 子表主食表 {product_name} sheet 写入数量失败{e}")
                    return
                # 往该没有内容的行中的G列写入金额
                try:
                    sheet.range((sub_row_index + 1, 9)).value = amount
                    print(f"Notice: 子表主食表 {product_name} sheet 写入金额成功")
                except Exception as e:
                    print(f"Error: 子表主食表 {product_name} sheet 写入金额失败{e}")
                    return
                
        except Exception as e:
            print(f"Error: {e}")
            return  


def update_sub_auxiliary_food_sheet(main_workbook, read_temp_storage_workbook, read_temp_storage_workbook_headers):
    """
    将暂存表数据提交到子表副食表
    Parameters:
     main_workbook: 子表副食表工作簿对象
     sub_auxiliary_food_excel_file_path: 子表副食表路径
     read_temp_storage_workbook: 暂存表格工作簿对象
     read_temp_storage_workbook_headers: 暂存表格表头
    
    :return: None
    """
    
    if not __main__.MODE:
        "子表副食表入库模式"
        try:

            # 轮询读取暂存表格数据行
            for row_index in range(1, read_temp_storage_workbook.sheet_by_index(0).nrows):
                # 读取行数据
                row_data = read_temp_storage_workbook.sheet_by_index(0).row_values(row_index)
                # 创建一个字典，用于存储列索引和列名的对应关系
                header_index = {name: idx for idx, name in enumerate(read_temp_storage_workbook_headers)}
                
                # 将日期分解为月和日
                year, month, day = row_data[header_index["日期"]].split("-")
                # 获取行中类别列类型单元中的类别名数据
                category_name = row_data[header_index["类别"]]
                # 获取行中品名列类型单元中的品名名数据
                product_name = row_data[header_index["品名"]]
                # 获取行中单位列类型单元中的单位名数据
                unit_name = row_data[header_index["单位"]]
                # 获取行中单价列类型单元中的单价名数据
                price = row_data[header_index["单价"]]
                # 获取行中数量列类型单元中的数量名数据
                quantity = row_data[header_index["数量"]]
                # 获取行中金额列单元中金额数据
                amount = row_data[header_index["金额"]]
                # 获取行中备注列单元中备注数据
                remark = row_data[header_index["备注"]]
                # 获取行中公司列单元中公司名数据
                company_name = row_data[header_index["公司"]]
                # 获取行中单名称列单元中单名数据
                single_name = row_data[header_index["单名"]]  

                if category_name == "副食":
                    # 获取所有sheet的name
                    sheet_names = [s.name for s in main_workbook.sheets]
                    # 筛选包含product_name的sheet名字
                    matching_sheets = [name for name in sheet_names if product_name in re.sub(r'\d+', '', name)]
                    print(matching_sheets)
                    # 取大于product_name长度且长度最小的sheet_name
                    if matching_sheets:
                        sheet_name = min((name for name in matching_sheets if len(re.sub(r'\d+', '', name)) >= len(product_name)), key=len, default=None)
                        if sheet_name:
                            sheet = main_workbook.sheets[sheet_name]
                        else:
                            print(f"Warning: 未找到合适的sheet匹配品名 {product_name}")
                            return
                    else:
                        print(f"Warning: 未在子表副食表中找到品名为 {product_name} 的表,已跳过本菜品的写入")
                        continue
                    print(f"Notice: 找到副食表sheet {sheet.name}")
                    #暂时感觉这个for循环没什么问题
                    #wjwcj: 2025/05/04 15:34
                    for sub_row_index in range(sheet.used_range.rows.count):
                        # 检查每行的1到11列是否都是空
                        if all(is_visually_empty(sheet.range((sub_row_index + 1, col))) for col in range(1, 12)):

                            # 向前检查是否是“过次页 + 空行 + 空行”的模式
                            if is_previous_rows_after_page_break(sheet, sub_row_index + 1):
                                print(f"Warning: 忽略第 {sub_row_index + 1} 行（前面是‘过次页’+连续空行）")
                                continue

                            print("Notice: 这里开始执行", str(sub_row_index + 1))   

                            # 检查前一行是否符合某些条件（仅包含空格或单个标点符号）
                            if sub_row_index > 0 and all(
                                ((sheet.range((sub_row_index, col)).value is None) or 
                                (is_single_punctuation(str(sheet.range((sub_row_index, col)).value).strip())))
                                for col in range(1, 12)
                            ):
                                print(f"Notice: 发现第 {sub_row_index + 1} 行可用(仅包含空格或单个标点)，开始写入数据")
                                break

                            print(f"Notice: 发现第 {sub_row_index + 1} 行为空行，开始写入数据")
                            break


                    # 往该没有内容的行的A列中写入月份、B列中写入日
                    try:
                        sheet.range((sub_row_index + 1, 1)).value = month
                        sheet.range((sub_row_index + 1, 2)).value = day
                        print(f"Notice: 子表副食表 {product_name} sheet 写入日期成功")
                    except Exception as e:
                        print(f"Error: 子表副食表 {product_name} sheet 写入日期失败{e}")
                        return
                    # 往该没有内容的行的D列中写入出入库摘要
                    try:
                        sheet.range((sub_row_index + 1, 4)).value = "入库"
                        print(f"Notice: 子表副食表 {product_name} sheet 写入出入库摘要成功")
                    except Exception as e:
                        print(f"Error: 子表副食表 {product_name} sheet 写入出入库摘要失败{e}")
                        return
                    # 往该没有内容的行中的E列写入单价
                    try:
                        sheet.range((sub_row_index + 1, 5)).value = price
                        print(f"Notice: 子表副食表 {product_name} sheet 写入单价成功")
                    except Exception as e:
                        print(f"Error: 子表副食表 {product_name} sheet 写入单价失败{e}")
                        return
                    # 往该没有内容的行中的F列写入数量
                    try:
                        sheet.range((sub_row_index + 1, 6)).value = quantity
                        print(f"Notice: 子表副食表 {product_name} sheet 写入数量成功")
                    except Exception as e:
                        print(f"Error: 子表副食表 {product_name} sheet 写入数量失败{e}")
                        return
                    # 往该没有内容的行中的G列写入金额
                    try:
                        sheet.range((sub_row_index + 1, 7)).value = amount
                        print(f"Notice: 子表副食表 {product_name} sheet 写入金额成功")
                    except Exception as e:
                        print(f"Error: 子表副食表 {product_name} sheet 写入金额失败{e}")
                        return
                
                else:
                    print(f"Warning: 在子表副食表入库时该食品类别属性不名为 主食 ,实名为 {category_name} 已跳过该菜品子表副食表写入")
                    continue
            
        except Exception as e:
            print(f"Error: 将暂存表数据提交到子表副食表出错，出错信息 {e}")
            return  

    else:
        "子表副食表入库模式"
        try:
            # 轮询读取暂存表格数据行
            for row_index in range(1, read_temp_storage_workbook.sheet_by_index(0).nrows):
                # 读取行数据
                row_data = read_temp_storage_workbook.sheet_by_index(0).row_values(row_index)
                # 创建一个字典，用于存储列索引和列名的对应关系
                header_index = {name: idx for idx, name in enumerate(read_temp_storage_workbook_headers)}
                
                try:
                    # 将日期分解为月和日
                    year, month, day = row_data[header_index["日期"]].split("-")
                    # 获取行中类别列类型单元中的类别名数据
                    category_name = row_data[header_index["类别"]]
                    # 获取行中品名列类型单元中的品名名数据
                    product_name = row_data[header_index["品名"]]
                    # 获取行中单位列类型单元中的单位名数据
                    unit_name = row_data[header_index["单位"]]
                    # 获取行中单价列类型单元中的单价名数据
                    price = row_data[header_index["单价"]]
                    # 获取行中数量列类型单元中的数量名数据
                    quantity = row_data[header_index["数量"]]
                    # 获取行中金额列单元中金额数据
                    amount = row_data[header_index["金额"]]
                    # 获取行中备注列单元中备注数据
                    remark = row_data[header_index["备注"]]
                    # 获取行中公司列单元中公司名数据
                    company_name = row_data[header_index["公司"]]
                    # 获取行中单名称列单元中单名数据
                    single_name = row_data[header_index["单名"]]  
                except Exception as e:
                    __main__.SAVE_OK_SIGNAL = False
                    print(f"Error: 将暂存表数据提交到子表副食表(出库)时拆解数据出错 {e}")


                # 获取所有sheet的name
                sheet_names = [s.name for s in main_workbook.sheets]
                # 筛选包含product_name的sheet名字
                matching_sheets = [name for name in sheet_names if product_name in re.sub(r'\d+', '', name)]
                print(matching_sheets)
                # 取大于product_name长度且长度最小的sheet_name
                if matching_sheets:
                    sheet_name = min((name for name in matching_sheets if len(re.sub(r'\d+', '', name)) >= len(product_name)), key=len, default=None)
                    if sheet_name:
                        sheet = main_workbook.sheets[sheet_name]
                    else:
                        print(f"Warning: 未找到合适的sheet匹配品名 {product_name}")
                        return
                else:
                    print(f"Warning: 未找到品名为 {product_name} 的sheet")
                    return
                
                for sub_row_index in range(sheet.used_range.rows.count):
                    # 检查每行的1到11列是否都是空
                    if all(is_visually_empty(sheet.range((sub_row_index + 1, col))) for col in range(1, 12)):

                        # 向前检查是否是“过次页 + 空行 + 空行”的模式
                        if is_previous_rows_after_page_break(sheet, sub_row_index + 1):
                            print(f"Warning: 忽略第 {sub_row_index + 1} 行（前面是‘过次页’+连续空行）")
                            continue

                        print("Notice: 这里开始执行", str(sub_row_index + 1))   

                        # 检查前一行是否符合某些条件（仅包含空格或单个标点符号）
                        if sub_row_index > 0 and all(
                            ((sheet.range((sub_row_index, col)).value is None) or 
                            (is_single_punctuation(str(sheet.range((sub_row_index, col)).value).strip())))
                            for col in range(1, 12)
                        ):
                            print(f"Notice: 发现第 {sub_row_index + 1} 行可用(仅包含空格或单个标点)，开始写入数据")
                            break

                        print(f"Notice: 发现第 {sub_row_index + 1} 行为空行，开始写入数据")
                        break


                # 往该没有内容的行的A列中写入月份、B列中写入日
                try:
                    sheet.range((sub_row_index + 1, 1)).value = month
                    sheet.range((sub_row_index + 1, 2)).value = day
                    print(f"Notice: 子表副食表 {product_name} sheet 写入日期成功")
                except Exception as e:
                    print(f"Error: 子表副食表 {product_name} sheet 写入日期失败{e}")
                    return
                # 往该没有内容的行的D列中写入出入库摘要
                try:
                    sheet.range((sub_row_index + 1, 4)).value = "出库"
                    print(f"Notice: 子表副食表 {product_name} sheet 写入出入库摘要成功")
                except Exception as e:
                    print(f"Error: 子表副食表 {product_name} sheet 写入出入库摘要失败{e}")
                    return
                # 往该没有内容的行中的E列写入单价
                try:
                    sheet.range((sub_row_index + 1, 5)).value = price
                    print(f"Notice: 子表副食表 {product_name} sheet 写入单价成功")
                except Exception as e:
                    print(f"Error: 子表副食表 {product_name} sheet 写入单价失败{e}")
                    return
                # 往该没有内容的行中的F列写入数量
                try:
                    sheet.range((sub_row_index + 1, 8)).value = quantity
                    print(f"Notice: 子表副食表 {product_name} sheet 写入数量成功")
                except Exception as e:
                    print(f"Error: 子表副食表 {product_name} sheet 写入数量失败{e}")
                    return
                # 往该没有内容的行中的G列写入金额
                try:
                    sheet.range((sub_row_index + 1, 9)).value = amount
                    print(f"Notice: 子表副食表 {product_name} sheet 写入金额成功")
                except Exception as e:
                    print(f"Error: 子表副食表 {product_name} sheet 写入金额失败{e}")
                    return
            
        except Exception as e:
            print(f"Error: {e}")
            return          


def update_welfare_food_sheet(self,app,welfare_food_excel_file_path,read_temp_storage_workbook,read_temp_storage_workbook_headers):
    """
    更新或添加数据到食堂福利表
    :param main_workbook: 主工作簿对象
    :param product_name: 物品名称
    :param unit_name: 计量单位
    :param quantity: 数量
    :param price: 单价
    :param amount: 金额
    :param remark: 备注
    :return: None
    """
    # 在福利表中更新信息
    try:
        # 读取福利表表格
        try:
            # 打开福利表簿对象
            main_workbook = app.books.open(welfare_food_excel_file_path)
            print(f"Notice: 福利表表加载成功，文件路径: {welfare_food_excel_file_path}")
        except Exception as e:
            __main__.SAVE_OK_SIGNAL = False
            print(f"Error: {e}")
            return    

        # 轮询读取暂存表格数据行
        for row_index in range(1, read_temp_storage_workbook.sheet_by_index(0).nrows):
            # 读取行数据
            row_data = read_temp_storage_workbook.sheet_by_index(0).row_values(row_index)
            # 创建一个字典，用于存储列索引和列名的对应关系
            header_index = {name: idx for idx, name in enumerate(read_temp_storage_workbook_headers)}
            
            # 将日期分解为月和日
            year, month, day = row_data[header_index["日期"]].split("-")
            # 获取行中类别列类型单元中的类别名数据
            category_name = row_data[header_index["类别"]]
            # 获取行中品名列类型单元中的品名名数据
            product_name = row_data[header_index["品名"]]
            # 获取行中单位列类型单元中的单位名数据
            unit_name = row_data[header_index["单位"]]
            # 获取行中单价列类型单元中的单价名数据
            price = row_data[header_index["单价"]]
            # 获取行中数量列类型单元中的数量名数据
            quantity = row_data[header_index["数量"]]
            # 获取行中金额列单元中金额数据
            amount = row_data[header_index["金额"]]
            # 获取行中备注列单元中备注数据
            remark = row_data[header_index["备注"]]
            # 获取行中公司列单元中公司名数据
            company_name = row_data[header_index["公司"]]
            # 获取行中单名称列单元中单名数据
            single_name = row_data[header_index["单名"]]

            if not __main__.MODE:
                print("Notice: 福利表正在入库")
                # 尝试打开名为过年福利入的 sheet
                if "过年福利入" in [s.name for s in main_workbook.sheets]:
                    sheet = main_workbook.sheets["过年福利入"]
                    print(f"Notice: 入库福利表时找到名为 `过年福利入` 的sheet")
                else:
                    print(f"Error: 入库福利表时未找到入库名为 `过年福利入` 的sheet")
                    return
                        
                # 从已有的行中查找第一行空行，记录下空行行标（从表格的第二行开始）
                for row_index in range(0, sheet.used_range.rows.count):
                    if sheet.range((row_index + 1, 1)).value is None and row_index != 0:
                        # 检查前一行是否包含“领导”二字
                        if row_index > 0:
                            previous_row_values = [
                            str(sheet.range((row_index, col)).value).strip()
                            for col in range(1, sheet.used_range.columns.count + 1)
                            if sheet.range((row_index, col)).value is not None
                            ]
                            if any("领导" in value for value in previous_row_values):
                                print(f"Notice: 第 {row_index} 行包含“领导”二字，继续查找下一行")
                                continue

                        # 检查当前列的前几行是否包含“序号”二字
                        column_values = [str(sheet.range((row, 1)).value).strip() for row in range(1, row_index + 1) if sheet.range((row, 1)).value is not None]
                        if not any("序号" in value for value in column_values):
                            print(f"Notice: 前 {row_index} 行未找到“序号”二字，继续查找下一行")
                            continue
                        break

                # 尝试写入一行数据
                try:
                    # 获取当前列中所有的序号值，排除空值并转换为整数
                    existing_numbers = []
                    for i in range(row_index):
                        box_value = sheet.range((i + 1, 1)).value
                        # 判定 box_value 是否为整型
                        try:
                            if box_value == '日计':
                                existing_numbers = [0]
                            else:
                                box_value = int(box_value)  # 尝试将值转换为整数
                                if str(box_value).isdigit():
                                    existing_numbers.append(box_value)  # 如果转换成功，则添加到列表中
                        except:
                            continue
                            
                    # 计算新的序号值
                    new_number = max(existing_numbers) + 1 if existing_numbers else 1
                    # 写入序号数据
                    sheet.range((row_index + 1, 1)).value = new_number
                    print(f"Notice: 在福利表表为入库类型 {single_name} 的第 {row_index} 行写入序号：{new_number} 成功")

                    # 为B、C列写入月份日期数据
                    sheet.range((row_index + 1, 2)).value = month
                    sheet.range((row_index + 1, 3)).value = day
                    print(f"Notice: 在福利表表为入库类型 {single_name} 的第 {row_index} 行写入月份：{month} 日：{day} 成功")

                    #动态获取表头行行数
                    name_row = 0
                    for i in range(6):
                        #在前六行里找吧
                        datas = [str(sheet.range((i + 1, col)).value).strip().replace(" ", "") for col in range(1, 12)]
                        if "单价" in datas and "数量" in datas and "金额" in datas:
                            name_row = i + 1
                            break

                    # 依次为D~K列写入数据(D、E列合并，需要加入跳过判断逻辑)
                    for col_index in range(4, 12):
                        if col_index == 5:
                            # 如果当前列是E列，则跳过
                            continue
                        else:
                            # 操作该单元时候，访问第该单元对应列的第四行单元获取该列的列名属性
                            # wjwcj：这可不一定，自购主食出库的列名属性就在第三行，所以得动态获取(到name_row)
                                
                            cell_attribute = sheet.range((name_row, col_index)).value

                            if isinstance(cell_attribute, str):
                                # 去除所有中文之间的空格
                                cell_attribute = re.sub(r'(?<=[\u4e00-\u9fa5])\s+(?=[\u4e00-\u9fa5])', '', cell_attribute)

                            print("Notice: 当前列", col_index, cell_attribute)
                            try:
                                if cell_attribute == "计量单位":
                                    # 如果该列名是单独的计量单位，手动匹配暂存表格中名为单位列的对应单元值
                                    sheet.range((row_index + 1, col_index)).value = unit_name
                                    print(f"Notice: 在福利表表为入库类型 {single_name} 的 {row_index} 行名为 {cell_attribute} 的列写入值 {row_data[header_index['单位']]} 成功")

                                else:
                                    # 在row_data中查找该列名对应的值，然后写入正在被操作的该单元中
                                    #print("正在写入" + cell_attribute + "  " + str(row_data[header_index[cell_attribute]]))
                                    if cell_attribute == "类别" and single_name  in ["自购主食入库等", "自购主食出库"]:
                                        row_data[header_index[cell_attribute]] = row_data[header_index[cell_attribute]] + single_name.strip("等").strip("自购主食")
                                    sheet.range((row_index + 1, col_index)).value = row_data[header_index[cell_attribute]]
                                    print(f"Notice: 在福利表表为入库类型 {single_name} 的 {row_index} 行名为 {cell_attribute} 的列写入值 {row_data[header_index[cell_attribute]]} 成功")

                            except KeyError:
                                __main__.SAVE_OK_SIGNAL = False
                                print(f"Error: 未在主表入/出库类型 {single_name} 找到名为 {cell_attribute} 的列")
                except Exception as e:
                    __main__.SAVE_OK_SIGNAL = False
                    print(f"Error: 出库时写入数据时发生错误 {e}")
            else:
                print("Notice: 正在出库")
                # 尝试打开名为过年福利入的 sheet
                if "过年福利出 (2)" in [s.name for s in main_workbook.sheets]:
                    sheet = main_workbook.sheets["过年福利出 (2)"]
                    print(f"Notice: 找到入库类型名为 `过年福利出 (2)` 的sheet")
                else:
                    __main__.SAVE_OK_SIGNAL = False
                    print(f"Error: 未找到入库类型名为 `过年福利出 (2)` 的sheet,可能存在空字符")
                    return
                        
                # 查找第一行空行，记录下空行行标（从表格的第二行开始）
                for row_index in range(0, sheet.used_range.rows.count):
                    if sheet.range((row_index + 1, 1)).value is None and row_index != 0:
                        # 检查前一行是否包含“领导”二字
                        if row_index > 0:
                            previous_row_values = [
                            str(sheet.range((row_index, col)).value).strip()
                            for col in range(1, sheet.used_range.columns.count + 1)
                            if sheet.range((row_index, col)).value is not None
                            ]
                            if any("领导" in value for value in previous_row_values):
                                print(f"Notice: 第 {row_index} 行包含“领导”二字，继续查找下一行")
                                continue

                        # 检查当前列的前几行是否包含“序号”二字
                        column_values = [
                            str(sheet.range((row, 1)).value).strip()
                            for row in range(1, row_index + 1)
                            if sheet.range((row, 1)).value is not None
                        ]
                        if not any("序号" in value for value in column_values):
                            print(f"Notice: 前 {row_index} 行未找到“序号”二字，继续查找下一行")
                            continue
                        break

                # 尝试写入一行数据
                try:
                    # 获取当前列中所有的序号值，排除空值并转换为整数
                    existing_numbers = []
                    for i in range(row_index):
                        box_value = sheet.range((i + 1, 1)).value
                        try:
                            if box_value == '日计':
                                existing_numbers = [0]
                            else:
                                box_value = int(box_value)  # 尝试将值转换为整数
                                if str(box_value).isdigit():
                                    existing_numbers.append(box_value)  # 如果转换成功，则添加到列表中
                        except:
                            continue
                    # 计算新的序号值
                    new_number = max(existing_numbers) + 1 if existing_numbers else 1
                    # 写入序号数据
                    sheet.range((row_index + 1, 1)).value = new_number
                    print(f"Notice: 在福利表表为出库类型 {single_name} 的第 {row_index} 行写入序号：{new_number} 成功")

                    # 为B、C列写入月份日期数据
                    sheet.range((row_index + 1, 2)).value = month
                    sheet.range((row_index + 1, 3)).value = day
                    print(f"Notice: 在福利表表为出库类型 {single_name} 的第 {row_index} 行写入月份：{month} 日：{day} 成功")

                    #动态获取表头行行数
                    name_row = 0
                    for i in range(6):
                        #在前六行里找吧
                        datas = [str(sheet.range((i + 1, col)).value).strip().replace(" ", "") for col in range(1, 12)]
                        if "单价" in datas and "数量" in datas and "金额" in datas:
                            name_row = i + 1
                            break

                    # 依次为D~K列写入数据(D、E列合并，需要加入跳过判断逻辑)
                    for col_index in range(4, 12):
                        if col_index == 5:
                            # 如果当前列是E列，则跳过
                            continue
                        else:
                            # 操作该单元时候，访问第该单元对应列的第四行单元获取该列的列名属性
                            # wjwcj：这可不一定，自购主食出库的列名属性就在第三行，所以得动态获取(到name_row)
                                
                            cell_attribute = sheet.range((name_row, col_index)).value

                            if isinstance(cell_attribute, str):
                                # 去除所有中文之间的空格
                                cell_attribute = re.sub(r'(?<=[\u4e00-\u9fa5])\s+(?=[\u4e00-\u9fa5])', '', cell_attribute)

                            print("Notice: 当前列", col_index, cell_attribute)
                            try:
                                if cell_attribute == "计量单位":
                                    # 如果该列名是单独的计量单位，手动匹配暂存表格中名为单位列的对应单元值
                                    sheet.range((row_index + 1, col_index)).value = unit_name
                                    print(f"Notice: 在福利表表为出库类型 {single_name} 的 {row_index} 行名为 {cell_attribute} 的列写入值 {row_data[header_index['单位']]} 成功")

                                else:
                                    # 在row_data中查找该列名对应的值，然后写入正在被操作的该单元中
                                    #print("正在写入" + cell_attribute + "  " + str(row_data[header_index[cell_attribute]]))
                                    if cell_attribute == "类别" and single_name  in ["自购主食入库等", "自购主食出库"]:
                                        row_data[header_index[cell_attribute]] = row_data[header_index[cell_attribute]] + single_name.strip("等").strip("自购主食")
                                    sheet.range((row_index + 1, col_index)).value = row_data[header_index[cell_attribute]]
                                    print(f"Notice: 在福利表表为出库类型 {single_name} 的 {row_index} 行名为 {cell_attribute} 的列写入值 {row_data[header_index[cell_attribute]]} 成功")

                            except KeyError:
                                __main__.SAVE_OK_SIGNAL = False
                                print(f"Error: 未在主表入/出库类型 {single_name} 找到名为 {cell_attribute} 的列")
                except Exception as e:
                    __main__.SAVE_OK_SIGNAL = False
                    print(f"Error: 出库时写入数据时发生错误 {e}")
            
        # 保存福利表工作簿 
        main_workbook.save()
        "如果开启页计功能，先不关闭工作簿"
        if not __main__.ADD_DAY_SUMMARY and not __main__.ADD_MONTH_SUMMARY and not __main__.ADD_PAGE_SUMMARY and not __main__.ADD_TOTAL_SUMMARY:
            # 关闭福利表工作簿
            main_workbook.close()

            
    except Exception as e:
            print(f"Error: 更新或添加数据到食堂福利表出错,出错信息{e}")
    

def export_update_inventory_sheet(main_workbook, product_name, unit_name, quantity, price, amount, remark):
    """
    更新或添加数据到食堂物品收发存库存表(出库)
    :param main_workbook: 主工作簿对象
    :param product_name: 物品名称
    :param unit_name: 计量单位
    :param quantity: 数量
    :param price: 单价
    :param amount: 金额
    :param remark: 备注
    :return: None
    """
    """
    例食堂物品收发存库存表
    2025/05/04 20:55 wjwcj测试没问题
    """
    # 尝试打开名为食堂物品收发库存表的 sheet
    if "食堂物品收发存库存表" in [s.name for s in main_workbook.sheets]:
        sheet = main_workbook.sheets["食堂物品收发存库存表"]
        print(f"Notice: 找到出库类型名为 `食堂物品收发存库存表` 的sheet")
    else:
        print(f"Error: 未找到出库类型名为 `食堂物品收发存库存表` 的sheet,可能存在空字符")
        return

    try:
        # 调用Excel VBA API 查找名为'名称'的 A列中是否存在该名称
        found = sheet.range("A:A").api.Find(product_name)
        if found is not None:
            # 如果存在，则更新该行的数据
            # 用遍历方式查找行索引，避免直接用 .row
            row_index = None
            for i in range(1, sheet.used_range.rows.count + 1):
                if sheet.range(f"A{i}").value == product_name:
                    row_index = i
                    break
            if row_index is None:
                print(f"Error: 在表 食堂物品收发存库存表 未找到名称为 {product_name} 的行")
                return
            print(f"Notice: 在表 食堂物品收发存库存表 找到名称为 {product_name} 的行,行号为{row_index}")

            # 判断quantity、price、amount的值是否为数值
            try:
                # 去掉金额字符中的空字符，防止出现类似 '10. 56' 的字符造成强制类型转换报错
                if isinstance(amount, str):
                    amount = amount.replace(" ", "")
                if isinstance(quantity, str):
                    quantity = quantity.replace(" ", "")
                if isinstance(price, str):
                    price = price.replace(" ", "")

                quantity = float(quantity)
                price = float(price)
                amount = float(amount)
            except:
                print(f"Error: quantity、price、amount的值必须为数值")
                return
            if isinstance(quantity, (int, float)) and isinstance(price, (int, float)) and isinstance(amount, (int, float)):
                # 在I列更新数量信息，J列更新单价信息，K列更新金额信息
                raw_value = []
                for alpha in "IJK":
                    if sheet.range(f"{alpha}{row_index}").value is None:
                        raw_value.append(0)
                    elif "." in str(sheet.range(f"{alpha}{row_index}").value):
                        raw_value.append(float(sheet.range(f"{alpha}{row_index}").value))
                    else:
                        raw_value.append(int(sheet.range(f"{alpha}{row_index}").value))
                sheet.range(f"I{row_index}").value = raw_value[0] + quantity
                sheet.range(f"J{row_index}").value = raw_value[1] + price
                sheet.range(f"K{row_index}").value = raw_value[2] + amount

                print(f"Notice: 在表 食堂物品收发存库存表 更新行信息 数量、单价、金额 的列,行号为{row_index}")
            else:
                print(f"Error: quantity、price、amount的值必须为数值")
                return
        else:
            print(f"Error: 在表 食堂物品收发存库存表 未找到名称为 {product_name} 的行, 直接退出 食堂物品收发存库存表 的出库函数")
            return

    except Exception as e:
        print(f"Error: 更新食堂物品收发存库存表时出错 {e}")

def export_update_receipt_storage_sheet(main_workbook, single_name, category_name, amount):
    """
    更新收发存表皮中的条目信息(出库)
    :param main_workbook: 主工作簿对象
    :param single_name: 单名信息
    :param category_name: 类别信息
    :param amount: 金额数据
    :return: None
    """
    """
    收发存表皮
    wjwcj: 2025/05/04 21:31 测试没问题
    """
    # 尝试打开名为收发存表皮的 sheet
    if "收发存表皮" in [s.name for s in main_workbook.sheets]:
        sheet = main_workbook.sheets["收发存表皮"]
        print(f"Notice: 找到出库类型名为 `收发存表皮` 的sheet")
    else:
        print(f"Error: 未找到出库类型名为 `收发存表皮` 的sheet,可能存在空字符")
        return

    # 提取输入数据的单名信息和类别信息进行行索引词匹配
    row_index_name = None
    if single_name == "扶贫主食出库":
        row_index_name = "主食（帮扶食品）"
    elif single_name == "扶贫副食出库":
        row_index_name = "副食（帮扶食品）"
    elif single_name == "自购主食出库":
        if category_name == "主食":
            row_index_name = "主食（自购）"
        elif category_name == "副食":
            row_index_name = "副食（自购）"
        else:
            print("Error: 自购主食出库 未找到类别信息，请检查输入数据")
    elif single_name == "场调面食出库":
        if category_name == "主食":
            row_index_name = "正常厂主食"
        elif category_name == "副食":
            row_index_name = "正常厂主食"
        else:
            print("Error: 场调面食出库 未找到类别信息，请检查输入数据")
    else:
        print("Error: 未找到出库类型信息，请检查输入数据")
        return

    # 调用Excel API用行索引名匹配行索引
    found_row = sheet.range("A:A").api.Find(row_index_name)
    if found_row is not None:
        try:
            found_row_index = sheet.range("A:A").value.index(row_index_name) + 1
            print(f"Notice: 在 收发存表皮 Sheet 中找到 {row_index_name} 的行索引为 {found_row_index}")
        except Exception as e:
            print(f"Error: 获取行索引出错 {e}")
            return
    else:
        print(f"Error: 在 收发存表皮 Sheet 中未找到 {row_index_name} 的行索引，请检查输入数据")
        return

    # 更新K列的金额数据
    if found_row_index:
        if sheet.range(found_row_index, 11).value is None:
            sheet.range(found_row_index, 11).value = float(amount)
            print(f"Notice: 发现 收发存表皮 Sheet   中 {row_index_name} 的金额数据不存在,现在添加数据为 {sheet.range(found_row_index, 11).value}")
        else:
            print(f"Notice: 在 收发存表皮 Sheet 中 {row_index_name} 的金额原始数据为 {sheet.range(found_row_index, 11).value}")
            sheet.range(found_row_index, 11).value = float(amount) + float(sheet.range(found_row_index, 11).value) # Fixed：修复了can only concatenate str (not "float") to str，遇到运算问题时尽可能的强制类型转换
            print(f"Notice: 在 收发存表皮 Sheet 中更新 {row_index_name} 的金额数据成功,现在数据为 {sheet.range(found_row_index, 11).value}")
    else:
        print(f"Error: 在 收发存表皮 Sheet 中更新 {row_index_name} 的金额数据失败，请检查输入数据")

def export_update_main_food_detail_sheet(main_workbook, single_name, category_name, amount):

    """
    更新主副食品明细账中的条目信息(出库)
    :param main_workbook: 主工作簿对象
    :param single_name: 单名信息
    :param category_name: 类别信息
    :param amount: 金额数据
    :return: None
    """

    """注意！！这个函数只负责主副食品明细账"""
    """
    主副食品明细账
    wjwcj: 2025/05/05 12:07 测试没问题
    """
    
    # 尝试打开名为主副食品明细账的 sheet
    if "主副食品明细账" in [s.name for s in main_workbook.sheets]:
        sheet = main_workbook.sheets["主副食品明细账"]
        print(f"Notice: 找到出库类型名为 `主副食品明细账` 的sheet")
    else:
        print(f"Error: 未找到出库类型名为 `主副食品明细账` 的sheet,可能存在空字符")
        return

    # 提取输入数据的单名信息和类别信息进行行列索引词匹配
    if single_name == "扶贫主食出库":
        row_index_name = "（帮扶食品）主副食"
        column_index_name = "主食出库"
    elif single_name == "扶贫副食出库":
        row_index_name = "（帮扶食品）主副食"
        column_index_name = "副食出库"
    elif single_name in "自购主食出库等":
        if category_name == "主食":
            row_index_name = "自购主副食"
            column_index_name = "主食出库"
        elif category_name == "副食":
            row_index_name = "自购主副食"
            column_index_name = "副食出库"
        else:
            print(f"Error: 查找 '自购主食出库' sheet时未找到对应的类别信息，请检查类别")
            return
    else:
        print(f"Error: 未找到出库类型名为 `自购主食出库等` 的sheet,可能存在空字符")
        return

    # 调用Excel API 进行行索引名匹配
    found_row = sheet.range("A:A").api.Find(row_index_name)
    if found_row is not None:
        try:
            found_row_index = sheet.range("A:A").value.index(row_index_name) + 1
            print(f"Notice: 在 主副食品明细账 Sheet中找到 {row_index_name} 的行索引为 {found_row_index}")
        except Exception as e:
            print(f"Error: 获取行索引出错 {e}")
            return
    else:
        print(f"Error: 在 主副食品明细账 Sheet中未找到 {row_index_name} 的行索引，请检查输入数据")
        return

    # 调用Excel API 进行列索引名匹配
    found_column = sheet.range("5:5").api.Find(column_index_name)
    if found_column is not None:
        try:
            found_column_index = found_column.Column
            print(f"Notice: 在 主副食品明细账 Sheet中找到 {column_index_name} 的列索引为 {found_column_index}")
        except Exception as e:
            print(f"Error: 获取列索引出错 {e}")
            return
    else:
        print(f"Error: 在 主副食品明细账 Sheet中未找到 {column_index_name} 的列索引，请检查输入数据")
        return

    # 更新相应单元的金额数据
    if found_column_index is not None and found_row_index is not None:
        cell_value = sheet.range(found_row_index, found_column_index).value
        if cell_value is None:
            sheet.range(found_row_index, found_column_index).value = float(amount)
            print(f"Notice: 在 主副食品明细账 Sheet中的 {row_index_name} {column_index_name} 的金额数据为空，已更新为 {amount}")
        else:
            print(f"Notice: 在 主副食品明细账 Sheet中的 {row_index_name} {column_index_name} 的原始金额数据为 {cell_value}")
            sheet.range(found_row_index, found_column_index).value = float(amount) + float(cell_value)
            print(f"Notice: 在 主副食品明细账 Sheet中的 {row_index_name} {column_index_name} 的现在金额数据为 {sheet.range(found_row_index, found_column_index).value}")



def add_counter(self, app, modle,main_excel_file_path, sub_main_food_excel_file_path, sub_auxiliary_food_excel_file_path,welfare_excel_file_path):
    """
    在主表和子表中添加日计、月计、页计、累计

    Parameters:
        main_excel_file_path: 主表路径
        sub_main_food_excel_file_path: 子表主食表路径
        sub_auxiliary_food_excel_file_path: 子表副食表路径
        welfare_excel_file_path: 福利表路径
        return: None
    """
    print("\n\n\nNotice: ", "开始添加日计\月计\页计\合计") 

    if not (__main__.ADD_DAY_SUMMARY or __main__.ADD_MONTH_SUMMARY or __main__.ADD_PAGE_SUMMARY or __main__.ADD_TOTAL_SUMMARY):    
        print("Warning: 没有添加日计、月计、页计或累计的选项，直接退出")
        self.worker.done2.emit()
        return

    try:
        
        if __main__.ONLY_WELFARE_TABLE == False:
            #添加主表
            note_main_table(self,app, main_excel_file_path)
            #添加子表主食表, wjwcj: 2025/05/13 12:43 测试没问题
            note_sub_main_table(self, app , modle,sub_main_food_excel_file_path)
            #添加子表副食表, wjwcj: 2025/05/13 12:43 测试没问题
            note_sub_auxiliary_table(self, app, modle,sub_auxiliary_food_excel_file_path)
        
        else:
            #添加福利表
            note_welfare_table(self, app,modle ,welfare_excel_file_path)

    except Exception as e:
        __main__.SAVE_OK_SIGNAL = False 
        print(f"Error:在主表和子表中添加日计、月计、页计、累计,出错信息 {e}")

def note_main_table(self, app ,  main_excel_file_path):
    """
    在主表中添加日计、月计、页计、合计
    :param main_excel_file_path 主表路径
    :return None
    """
    
    print("\nNotice: ", "开始添加主表日计\月计\页计\合计")

    workbook = None

    #主表：各种杂项需要做日计月计, "日计"放"序号"--金额
    if __main__.ADD_DAY_SUMMARY or __main__.ADD_MONTH_SUMMARY or __main__.ADD_PAGE_SUMMARY or __main__.ADD_TOTAL_SUMMARY:
        sheets_to_add = get_all_sheets_todo_for_main_table()
    else:
        return 
    
    # 继承之前已经打开的工作簿对象
    for wb in app.books:
        # 获取路径的文件名
        excel_name = os.path.basename(main_excel_file_path)
        if wb.name == excel_name:
            workbook = wb
            break

    "添加日记"
    if __main__.ADD_DAY_SUMMARY:
        # 在临时储存表格7里查找单名，只有有的单名需要添加
        # 同时查找手动和图片的两个临时储存表，排除重复项(用{}即可去重)
        # 前文获取了单名
        
        print("Notice: ", "开始添加主表日计")

        for sheet_name in sheets_to_add:
            
            # 寻找月份和日期都匹配的行数matching_rows
            matching_rows = find_matching_today_rows(main_excel_file_path, sheet_name=sheet_name)
            print("Notice: 重复行", matching_rows)
            print(sheet_name)
            try:
                sheet = workbook.sheets[sheet_name]  # 使用指定的工作表名称
            except:
                print("sheet名不存在")
                continue
            total_amount = 0
            
            for row in matching_rows:
                total_amount += round(float(sheet.range((row, 10)).value), 2)
            # 求出总金额后加一行
            row_index = find_the_first_empty_line_in_main_excel(sheet)
            print(f"在{row_index + 1}写入日计")
            try:
                sheet.range((row_index + 1, 1)).value = "日计"  # 在A列写入“日计”
                sheet.range((row_index + 1, 10)).value = total_amount  # 在J列写入总金额
                print("Notice: ", f"主表sheet {sheet_name} 日计添加完成")
                print(f"Notice: 主表文件 {main_excel_file_path} 保存完成")   

            except Exception as e:
                print(f"Error: 无法写入日计数据到主表sheet {sheet_name}, 错误信息: {e}")
       
        print("Notice: ", "主表日计全部添加完成")

    "添加月计"
    if __main__.ADD_MONTH_SUMMARY:

        print("Notice: ", "开始添加主表月计")

        # 在临时储存表格里查找单名，只有有的单名需要添加
        # 同时查找手动和图片的两个临时储存表，排除重复项(用{}即可去重)
        # 前文获取了单名
        
        for sheet_name in sheets_to_add:
            #寻找月份和日期都匹配的行数matching_rows
            matching_rows = find_matching_month_rows(main_excel_file_path, sheet_name=sheet_name)
            print("重复行", matching_rows)

            try:
                sheet = workbook.sheets[sheet_name]  # 使用指定的工作表名称
            except:
                print("Warning:sheet名不存在")
                continue
            total_amount = 0
            for row in matching_rows:
                total_amount += round(float(sheet.range((row, 10)).value), 2)
            # 求出总金额后加一行
            row_index = find_the_first_empty_line_in_main_excel(sheet)
            print(f"在{row_index + 1}写入月计")
            try:
                sheet.range((row_index + 1, 1)).value = "月计"  # 在A列写入“月计”
                sheet.range((row_index + 1, 10)).value = total_amount  # 在J列写入总金额
                workbook.save()
                print("Notice: ", f"主表sheet {sheet_name} 月计添加完成")
            except Exception as e:
                print(f"Error: 无法写入月计数据到主表sheet {sheet_name}, 错误信息: {e}")
        
        print("Notice: ", "主表月计全部添加完成")


    "添加页计"
    if __main__.ADD_PAGE_SUMMARY:
      
        print("Notice: ", "开始添加主表页计")

        for sheet_name in sheets_to_add:
                      
            # 获取 workbook 对象中所有表的表名
            sheet_names = [s.name for s in workbook.sheets]
            # 检查 sheet_name 是否在 sheet_names 中
            for name in sheet_names:
                # 去除空格
                if sheet_name == re.sub(r'\s+', '', name):
                    sheet_name = name
                    break
                
            try:
                work_sheet = workbook.sheets[sheet_name]  # 使用指定的工作表名称
                counting_page_value("主表",workbook,work_sheet)
            
            except Exception as e:  
                print(f"Error:为主表 {sheet_name} sheet 添加页计时报错，错误信息: {e}")
                __main__.SAVE_OK_SIGNAL = False 
                continue
           
        print("Notice: ", "主表页计全部添加完成")
    
    workbook.save()  # 保存工作簿
    workbook.close()  # 关闭工作簿
    print("Notice: ", "主表文件保存完成")
    

    



def note_sub_main_table(self, app,model,sub_main_food_excel_file_path):
    """
    在子主食表添加日计、月计、页计、合计
    Parameters:
     self: 主窗口对象
     app: Excel应用程序对象
     model: 模式
     sub_main_food_excel_file_path: 子主食表路径
    :return None
    """

    print("\nNotice: ", "开始添加子表主食表日计\月计\页计\合计")

    workbook = None
    #在暂存的表里面查找第二列"品名", 将其作为sheet名查找对应sheet
    if __main__.ADD_DAY_SUMMARY or __main__.ADD_MONTH_SUMMARY or __main__.ADD_PAGE_SUMMARY or __main__.ADD_TOTAL_SUMMARY:
        sheets_to_add = get_all_sheets_todo_for_sub_table(app,model)
    else:
        return 
    
    # 继承之前已经打开的工作簿对象
    for wb in app.books:
        # 获取路径的文件名
        excel_name = os.path.basename(sub_main_food_excel_file_path)
        if wb.name == excel_name:
            workbook = wb
            break
    
    if __main__.ADD_DAY_SUMMARY:

        print("Notice: ", "开始添加子表主食表日计")

        for product_name in sheets_to_add:

            #这里查找正确的sheet名

            sheet_names = [s.name for s in workbook.sheets]
            # 筛选包含product_name的sheet名字
            matching_sheets = [name for name in sheet_names if product_name in re.sub(r'\d+', '', name)]
            print(matching_sheets)
            # 取大于product_name长度且长度最小的sheet_name
            if matching_sheets:
                sheet_name = min((name for name in matching_sheets if len(re.sub(r'\d+', '', name)) >= len(product_name)), key=len, default=None)
                if sheet_name:
                    try:
                        sheet = workbook.sheets[sheet_name]
                    except:
                        print("sheet名不存在")
                        continue
                else:
                    print(f"未找到合适的sheet匹配品名 {product_name}")
                    return
            else:
                print(f"Warning: 未找到品名为 {product_name} 的sheet")
                return
            #然后开始写入日计/月计
            matching_rows = find_matching_today_rows(sub_main_food_excel_file_path, sheet_name=sheet_name, columns=[1, 2])
            print("重复行", matching_rows)
            try:
                sheet = workbook.sheets[sheet]  # 使用指定的工作表名称
            except:
                print("sheet名不存在")
                continue
            in_quantity = 0
            in_amount = 0
            out_quantity = 0
            out_amount = 0
            for row in matching_rows:
                in_quantity += round(float(sheet.range((row, 6)).value) if sheet.range((row, 6)).value is not None else 0, 2)  # F列
                in_amount += round(float(sheet.range((row, 7)).value) if sheet.range((row, 7)).value is not None else 0, 2)    # G列
                out_quantity += round(float(sheet.range((row, 8)).value) if sheet.range((row, 8)).value is not None else 0, 2)  # H列
                out_amount += round(float(sheet.range((row, 9)).value) if sheet.range((row, 9)).value is not None else 0, 2)    # I列
            # 求出总金额后加一行
            row_index = find_the_first_empty_line_in_sub_main_excel(sheet)
            print(f"在{row_index}写入日计")
            
            try:
                #这里也死活写不进去
                sheet.range((row_index, 4)).value = "日计"  # 在D列写入“日计”
                data = [in_quantity, in_amount, out_quantity, out_amount]
                for i in range(len(data)):
                    sheet.range((row_index, 6 + i)).value = data[i]
                #抄写库存数量与金额
                sheet.range((row_index, 10)).value = sheet.range((row_index - 1, 10)).value
                sheet.range((row_index, 11)).value = sheet.range((row_index - 1, 11)).value

                print("Notice: ", f"主表sheet {sheet_name} 日计添加完成")
            except Exception as e:
                print(f"Error: 无法写入日计数据到子表sheet {sheet_name}, 错误信息: {e}")
            
        print("Notice: ", "子表主食表日计全部添加完成")
    
    if __main__.ADD_MONTH_SUMMARY:

        print("Notice: ", "开始添加子表主食表月计")

        for product_name in sheets_to_add:

            print(sub_main_food_excel_file_path)
            
            sheet_names = [s.name for s in workbook.sheets]
            # 筛选包含product_name的sheet名字
            matching_sheets = [name for name in sheet_names if product_name in re.sub(r'\d+', '', name)]
            print(matching_sheets)
            # 取大于product_name长度且长度最小的sheet_name
            if matching_sheets:
                sheet_name = min((name for name in matching_sheets if len(re.sub(r'\d+', '', name)) >= len(product_name)), key=len, default=None)
                if sheet_name:
                    sheet = workbook.sheets[sheet_name]
                else:
                    print(f"未找到合适的sheet匹配品名 {product_name}")
                    return
            else:
                print(f"Warning: 未找到品名为 {product_name} 的sheet")
                return
            #然后开始写入月计
            matching_rows = find_matching_month_rows(sub_main_food_excel_file_path, sheet_name=sheet_name, columns=[1, 2])
            print("重复行", matching_rows)
            try:
                sheet = workbook.sheets[sheet]  # 使用指定的工作表名称
            except:
                print("sheet名不存在")
                continue
            in_quantity = 0
            in_amount = 0
            out_quantity = 0
            out_amount = 0
            for row in matching_rows:
                in_quantity += round(float(sheet.range((row, 6)).value) if sheet.range((row, 6)).value is not None else 0, 2)  # F列
                in_amount += round(float(sheet.range((row, 7)).value) if sheet.range((row, 7)).value is not None else 0, 2)    # G列
                out_quantity += round(float(sheet.range((row, 8)).value) if sheet.range((row, 8)).value is not None else 0, 2)  # H列
                out_amount += round(float(sheet.range((row, 9)).value) if sheet.range((row, 9)).value is not None else 0, 2)    # I列
            # 求出总金额后加一行
            row_index = find_the_first_empty_line_in_sub_main_excel(sheet)
            print(f"在{row_index}写入月计")
            try:
                #这里也死活写不进去
                sheet.range((row_index, 4)).value = "月计"  # 在D列写入“月计”
                data = [in_quantity, in_amount, out_quantity, out_amount]
                for i in range(len(data)):
                    sheet.range((row_index, 6 + i)).value = data[i]
                #抄写库存数量与金额
                sheet.range((row_index, 10)).value = sheet.range((row_index - 1, 10)).value
                sheet.range((row_index, 11)).value = sheet.range((row_index - 1, 11)).value

                print("Notice: ", f"主表sheet {sheet_name} 月计添加完成")
            except Exception as e:
                print(f"Error: 无法写入月计数据到子表sheet {sheet_name}, 错误信息: {e}")
        print("NOtice: ", "子表月计全部添加完成")

        "添加页计"
        if __main__.ADD_PAGE_SUMMARY:
            
            print("Notice: ", "开始添加子表主食表页计")

            for sheet_name in sheets_to_add:
                          
                # 获取 workbook 对象中所有表的表名
                sheet_names = [s.name for s in workbook.sheets]
                for name in sheet_names:
                    # 去除空格
                    if sheet_name == re.sub(r'\s+', '', name):
                        sheet_name = name
                        break
                # 尝试获取工作表对象    
                try:

                    work_sheet = workbook.sheets[sheet_name]  # 使用指定的工作表名称
                    counting_page_value("子表主食表",workbook,work_sheet)

                except Exception as e:  
                    print(f"Error:为子表主食表 {sheet_name} sheet 添加页计时报错，错误信息: {e}")
                    __main__.SAVE_OK_SIGNAL = False 
                    continue

            print("Notice: ", "子表主食表页计全部添加完成")   

    workbook.save()  # 保存工作簿
    workbook.close()  # 关闭工作簿
    print("Notice: ", "添加子表主食表日计\月计\页计\合计完成")
    



def note_sub_auxiliary_table(self,app, model,sub_auxiliary_food_excel_file_path):
    """
    在子表副食表添加日计、月计、页计、总计
    
    Parameters:
        app: Excel应用程序对象
        model: 模式
        sub_auxiliary_food_excel_file_path: 子副食表路径

    :return: None
    """
    print("\nNotice: ", "开始添加子表副食表日计\月计\页计\总计")

    workbook = None
    #在暂存的表里面查找"品名", 将其作为sheet名查找对应sheet
    if __main__.ADD_DAY_SUMMARY or __main__.ADD_MONTH_SUMMARY or __main__.ADD_PAGE_SUMMARY or __main__.ADD_TOTAL_SUMMARY:
        sheets_to_add = get_all_sheets_todo_for_sub_table(app,model)
    else:
        return 
    
    # 继承之前已经打开的工作簿对象
    for wb in app.books:
        # 获取路径的文件名
        excel_name = os.path.basename(sub_auxiliary_food_excel_file_path)
        if wb.name == excel_name:
            workbook = wb
            break

    if __main__.ADD_DAY_SUMMARY:
        for product_name in sheets_to_add:
            #这里查找正确的sheet名
            sheet_name = ""

            
            sheet_names = [s.name for s in workbook.sheets]
            # 筛选包含product_name的sheet名字
            matching_sheets = [name for name in sheet_names if product_name in re.sub(r'\d+', '', name)]
            print(matching_sheets)
            # 取大于product_name长度且长度最小的sheet_name
            if matching_sheets:
                sheet_name = min((name for name in matching_sheets if len(re.sub(r'\d+', '', name)) >= len(product_name)), key=len, default=None)
                if sheet_name:
                    sheet = workbook.sheets[sheet_name]
                else:
                    print(f"未找到合适的sheet匹配品名 {product_name}")
                    return
            else:
                print(f"Warning: 未找到品名为 {product_name} 的sheet")
                return
            #然后开始写入日计/月计
            matching_rows = find_matching_today_rows(sub_auxiliary_food_excel_file_path, sheet_name=sheet_name, columns=[1, 2])
            print("重复行", matching_rows)
            try:
                sheet = workbook.sheets[sheet]  # 使用指定的工作表名称
            except:
                print("sheet名不存在")
                continue
            in_quantity = 0
            in_amount = 0
            out_quantity = 0
            out_amount = 0
            for row in matching_rows:
                in_quantity += (round(float(sheet.range((row, 6)).value), 2) if sheet.range((row, 6)).value is not None else 0)  # F列
                in_amount += (round(float(sheet.range((row, 7)).value), 2) if sheet.range((row, 7)).value is not None else 0)    # G列
                out_quantity += (round(float(sheet.range((row, 8)).value), 2) if sheet.range((row, 8)).value is not None else 0)  # H列
                out_amount += (round(float(sheet.range((row, 9)).value), 2) if sheet.range((row, 9)).value is not None else 0)    # I列
            # 求出总金额后加一行(这个函数里已经加了)
            row_index = find_the_first_empty_line_in_sub_auxiliary_excel(sheet)
            print(f"在{row_index}写入日计")
            try:
                sheet = workbook.sheets[sheet_name]
                print(f"在{sheet_name}写入日计")
                sheet.range((row_index, 4)).value = "日计"  # 在D列写入“日计”
                data = [in_quantity, in_amount, out_quantity, out_amount]
                for i in range(len(data)):
                    sheet.range((row_index, 6 + i)).value = data[i]
                #抄写库存数量与金额
                sheet.range((row_index, 10)).value = sheet.range((row_index - 1, 10)).value
                sheet.range((row_index, 11)).value = sheet.range((row_index - 1, 11)).value
      
                print("Notice: ", f"子表副食表sheet {sheet_name} 日计添加完成")
            except Exception as e:
                print(f"Error: 无法写入日计数据到子表副食表sheet {sheet_name}, 错误信息: {e}")
            print("NOtice: ", "子表副食表日计全部添加完成")

    if __main__.ADD_MONTH_SUMMARY:
        for product_name in sheets_to_add:

            print(sub_auxiliary_food_excel_file_path)
            
            sheet_names = [s.name for s in workbook.sheets]
            # 筛选包含product_name的sheet名字
            matching_sheets = [name for name in sheet_names if product_name in re.sub(r'\d+', '', name)]
            print(matching_sheets)
            # 取大于product_name长度且长度最小的sheet_name
            if matching_sheets:
                sheet_name = min((name for name in matching_sheets if len(re.sub(r'\d+', '', name)) >= len(product_name)), key=len, default=None)
                if sheet_name:
                    sheet = workbook.sheets[sheet_name]
                else:
                    print(f"未找到合适的sheet匹配品名 {product_name}")
                    return
            else:
                print(f"Warning: 未找到品名为 {product_name} 的sheet")
                return
            #然后开始写入月计
            matching_rows = find_matching_month_rows(sub_auxiliary_food_excel_file_path, sheet_name=sheet_name, columns=[1, 2])
            print("重复行", matching_rows)
            try:
                sheet = workbook.sheets[sheet]  # 使用指定的工作表名称
            except:
                print("sheet名不存在")
                continue
            in_quantity = 0
            in_amount = 0
            out_quantity = 0
            out_amount = 0
            for row in matching_rows:
                in_quantity += round(float(sheet.range((row, 6)).value) if sheet.range((row, 6)).value is not None else 0, 2)  # F列
                in_amount += round(float(sheet.range((row, 7)).value) if sheet.range((row, 7)).value is not None else 0, 2)    # G列
                out_quantity += round(float(sheet.range((row, 8)).value) if sheet.range((row, 8)).value is not None else 0, 2)  # H列
                out_amount += round(float(sheet.range((row, 9)).value) if sheet.range((row, 9)).value is not None else 0, 2)    # I列
            # 求出总金额后加一行
            row_index = find_the_first_empty_line_in_sub_auxiliary_excel(sheet)
            print(f"在{row_index}写入月计")
            try:
                #这里也死活写不进去
                sheet.range((row_index, 4)).value = "月计"  # 在D列写入“月计”
                data = [in_quantity, in_amount, out_quantity, out_amount]
                for i in range(len(data)):
                    sheet.range((row_index, 6 + i)).value = data[i]
                #抄写库存数量与金额
                sheet.range((row_index, 10)).value = sheet.range((row_index - 1, 10)).value
                sheet.range((row_index, 11)).value = sheet.range((row_index - 1, 11)).value

                print("Notice: ", f"子副食表sheet {sheet_name} 月计添加完成")
            except Exception as e:
                print(f"Error: 无法写入月计数据到子副食表sheet {sheet_name}, 错误信息: {e}")
        print("NOtice: ", "子副食表月计全部添加完成")
        
        "添加页计"
        if __main__.ADD_PAGE_SUMMARY:
        
            for sheet_name in sheets_to_add:
                workbook = app.books.open(sub_auxiliary_food_excel_file_path)
                counting_page_value("子表副食表",workbook,sheet,sheet_name)

            print("Notice: ", "子表副食表页计全部添加完成")   


    workbook.save()  # 保存工作簿
    workbook.close()  # 关闭工作簿
    print("Notice: ", "添加子表副食表日计\月计\页计\总计完成")



def note_welfare_table(self, app,modle,welfare_excel_file_path):
    """
    为福利表增加日记、月计、页计、合计等数据
    Parameters:
        self: 主窗口对象
        app: Excel应用程序对象
        modle: 模式
        welfare_excel_file_path: 福利表路径
    :return: None
    """
    print("\nNotice: ", "开始添加福利表日计\月计\页计\合计")

    workbook = None

    #主表：各种杂项需要做日计月计, "日计"放"序号"--金额
    if __main__.ADD_DAY_SUMMARY or __main__.ADD_MONTH_SUMMARY or __main__.ADD_PAGE_SUMMARY or __main__.ADD_TOTAL_SUMMARY:
        #事实上这个和主表可以共用一个函数
        sheets_to_add = get_all_sheets_todo_for_main_table()
    else:
        return 
    
    # 继承之前已经打开的工作簿对象
    for wb in app.books:
        # 获取路径的文件名
        excel_name = os.path.basename(welfare_excel_file_path)
        if wb.name == excel_name:
            workbook = wb
            break

    if __main__.ADD_DAY_SUMMARY:
        for product_name in sheets_to_add:
            #这里查找正确的sheet名
            sheet_name = ""
            
            sheet_names = [s.name for s in workbook.sheets]
            for sheet_name in sheet_names:
                if sheet_name == product_name:
                    break
            sheet = workbook.sheets[sheet_name]
            print(f"待匹配{product_name}", f"匹配到{sheet_name}")
            
            matching_rows = find_matching_today_rows(welfare_excel_file_path, sheet_name=sheet_name, columns=[2, 3])
            print("重复行", matching_rows)
            try:
                sheet = workbook.sheets[sheet]  # 使用指定的工作表名称
            except:
                print("福利表sheetname不存在")
            quantity = 0
            amount = 0
            for row in matching_rows:
                quantity += (round(float(sheet.range((row, 9)).value), 2) if sheet.range((row, 9)).value is not None else 0)  # I列
                amount += (round(float(sheet.range((row, 10)).value), 2) if sheet.range((row, 10)).value is not None else 0)    # J列
            # 求出总金额后加一行(这个函数里已经加了)
            row_index = find_the_first_empty_line_in_main_excel(sheet) + 1
            print(f"在{row_index}写入日计")
            try:
                sheet = workbook.sheets[sheet_name]
                print(f"在{sheet_name}写入日计")
                sheet.range((row_index, 1)).value = "日计"  # 在A列写入“日计”
                data = [quantity, amount]
                for i in range(len(data)):
                    sheet.range((row_index, 9 + i)).value = data[i]

                print("Notice: ", f"福利表sheet {sheet_name} 日计添加完成")
            except Exception as e:
                print(f"Error: 无法写入日计数据到福利表sheet {sheet_name}, 错误信息: {e}")
            print("NOtice: ", "福利表日计全部添加完成")
    
    if __main__.ADD_MONTH_SUMMARY:
        for product_name in sheets_to_add:


            sheet_names = [s.name for s in workbook.sheets]
            for sheet_name in sheet_names:
                if sheet_name == product_name:
                    break
            sheet = workbook.sheets[sheet_name]
            print(f"待匹配{product_name}", f"匹配到{sheet_name}")
            matching_rows = find_matching_month_rows(welfare_excel_file_path, sheet_name=sheet_name, columns=[2, 3])
            print("重复行", matching_rows)
            sheet = workbook.sheets[sheet]  # 使用指定的工作表名称
            quantity = 0
            amount = 0
            for row in matching_rows:
                quantity += (round(float(sheet.range((row, 9)).value), 2) if sheet.range((row, 9)).value is not None else 0)  # I列
                amount += (round(float(sheet.range((row, 10)).value), 2) if sheet.range((row, 10)).value is not None else 0)    # J列
            # 求出总金额后加一行(这个函数里已经加了)
            row_index = find_the_first_empty_line_in_main_excel(sheet) + 1
            print(f"在{row_index}写入月计")
            try:
                sheet = workbook.sheets[sheet_name]
                print(f"在{sheet_name}写入月计")
                sheet.range((row_index, 1)).value = "月计"  # 在A列写入“月计”
                data = [quantity, amount]
                for i in range(len(data)):
                    sheet.range((row_index, 9 + i)).value = data[i]
                print("Notice: ", f"福利表sheet {sheet_name} 月计添加完成")
            except Exception as e:
                print(f"Error: 无法写入月计数据到福利表sheet {sheet_name}, 错误信息: {e}")
            print("NOtice: ", "福利表月计全部添加完成")

    "添加页计"
    if __main__.ADD_PAGE_SUMMARY:  
            
        for sheet_name in sheets_to_add:
            workbook = app.books.open(welfare_excel_file_path)
            counting_page_value("福利表",workbook,sheet,sheet_name)

        print("Notice: ", "福利表页计全部添加完成")   
    
    workbook.save()  # 保存工作簿
    workbook.close()  # 关闭工作簿
    print("Notice: ", "添加福利表日计\月计\页计\合计完成")


def img_excel_after_process(self,img_to_excel_file_path:str = os.path.abspath("./src/data/input/manual/temp_img_input.xlsx")):
    # 弹窗提示表格初步转录完成
    self.pushButton_4.setText("暂存该条")
    self.reply = QMessageBox.information(None, "提示", "图片转表格完成", QMessageBox.Ok | QMessageBox.Cancel)
    if self.reply == QMessageBox.Ok:

        # 调用 openpyxl 读取 Excel 文件
        try:
            workbook = openpyxl.load_workbook(img_to_excel_file_path)
            sheet = workbook.active
            # 将B2单元覆写为品名
            sheet['B2'] = '品名'
            # 将第二行第一列单元覆写成日期文本
            sheet['A2'] = '日期'
            # 为G2列加上备注
            sheet['G2'] = '备注'
            # 为H2列加上公司
            sheet['H2'] = '公司'
            # 为I2列加上单名
            sheet['I2'] = '单名'
            # 在B列前插入一列，并且在该列的第二行中插入'类别'
            sheet.insert_cols(2)
            sheet.cell(row=2, column=2).value = '类别'
            # 删除第一行
            sheet.delete_rows(1)
            # 保存并关闭文件
            workbook.save(img_to_excel_file_path)
            workbook.close()
            print("Notice: 图片转录表格后处理完成")

        except Exception as e:
            print(f"Error:后处理 {img_to_excel_file_path} 失败,错误信息 {e}")
    
    
        


    
# Summerize：
# 1. 索引类操作一定要考虑容错机制
# 2. 断点不要打在 return 上，否则执行 step over 根本不会执行 
# Learning:
# 1. Openpyxl 不对 .xls 文件格式提供支持，只能对 .xlsx 文件格式提供支持
# 2. 代码操作Excel打开的表时候会出现权限遭拒错误
# 3. xlrd&xlwd 和 xlrd&xlutils 两种库的搭配对于 Excel xls的表格操作容易触发兼容性问题
# 4. 
# TODO:
# - [x] 修复数据存储到Excel文件中的报错:ValueError: If using all scalar values, you must pass an index
# - [X] 实现以相对路径的方式存储表格到指定目录
# - [x] 2025.4.30 实现追加写入表格的行逻辑
# - [ ] 2025.5.1 实现数据提交到主表、副表Excel文件的功能
#   - [x] 修复Error: openpyxl does not support the old .xls file format, please use xlrd to read this file, or convert it to the more recent .xlsx file format.
#   - [x] 修复NameError: name 'input_data' is not defined
#   - [x] 实现提交条目数据到主表中
#      - [x] 将openpyxl替换为xlwd，实现Excel以xls文件保存，减少与原表格的数据格式冲突
#      - [x] 修复： [Errno 13] Permission denied: '.\\src\\data\\input\\manual\\temp_manual_input_data.xls'
#      - [x] 修复表单访问方法错用的问题
#      - [x] 修复Error: 'Worksheet' object has no attribute 'cell'
#      - [x] 修复TypeError: descriptor 'decode' for 'bytes' objects doesn't apply to a 'NoneType' object
#      - [x] 实现提交数据条目到主表物品相应的入库类型sheet表中
#      - [x] 实现提交数据条目到食堂物品收发存表中
#      - [x] 2025.5.3. 实现提交数据条目到主副食明细账中
#      - [x] 2025.5.3. 实现提交数据条目到收发表存皮重
#   - [x] 2025.5.3. 实现提交数据到子表中
#      - [x] 实现提交数据到主食表入库中
#      - [x] 实现提交数据到副食表入库中 
# - [x] 2025.5.1 修复暂存一次表格前7行出现None字符的问题
# - [x] 2025.5.2 解决store_single_entry_to_temple_excel函数表格不存在时[Errno 2] No such file or directory: '.\\src\\data\\input\\manual\\temp_manual_input_data.xls'的问题